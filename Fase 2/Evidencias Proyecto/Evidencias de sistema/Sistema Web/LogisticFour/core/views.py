# =============================================
#  LIBRERÍAS ESTÁNDAR DE PYTHON
# =============================================
import csv
from io import BytesIO
import requests
import csv
import json
from django.views.decorators.http import require_http_methods

from decimal import Decimal
from django.db import transaction
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from .forms import OrdenCompraForm, FacturaProveedorForm, RecepcionMercaderiaForm, ProductoForm
import requests
from django.http import JsonResponse
from django.views.decorators.http import require_GET


from .utils import ensure_ubicacion_sucursal

import logging
from datetime import datetime, timezone


from decimal import Decimal, InvalidOperation
from django.db.models.signals import pre_save, post_save
from django.dispatch import receiver
from django.utils import timezone
# core/views.py  (arriba con el resto de imports)
from django.shortcuts import render, redirect

# =============================================
#  LIBRERÍAS DE DJANGO
# =============================================
from django import forms
from django.apps import apps
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User, Group
from django.contrib.messages.views import SuccessMessageMixin
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import (
    Q, Sum, F, Value, Count, DecimalField, ExpressionWrapper
)
from django.db.models.functions import Coalesce, Lower
from django.http import (
    HttpResponse, Http404, JsonResponse,
    HttpResponseBadRequest, HttpResponseRedirect
)
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy, NoReverseMatch
from django.views import View
from django.views.decorators.http import require_POST, require_GET
from django.views.generic import (
    ListView, CreateView, UpdateView,
    DeleteView, DetailView
)

# =============================================
#  MÓDULOS DEL PROYECTO (CORE)
# =============================================
from core.forms import *
from core.forms import SignupUserForm, UsuarioPerfilForm
from core.models import *
from core.utils import ensure_ubicacion_sucursal, qr_url, barcode_url
from django.db import models as djmodels

from core.forms import SignupUserForm, UsuarioPerfilForm, FinanzasReporteForm
from core.models import UsuarioPerfil, OrdenCompra, RecepcionMercaderia, FacturaProveedor, UsuarioPerfil, Producto, Stock
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string

def notificar_stock_bajo(producto, nombre_lugar, stock_actual):
    """
    Envía correo HTML cuando el stock en 'nombre_lugar' baja de 10.
    """
    if stock_actual is None or stock_actual >= 10:
        return

    context = {
        "producto": producto,
        "sku": producto.sku,
        "sucursal": nombre_lugar,   # puede ser sucursal o bodega
        "stock": stock_actual,
    }

    subject = f"⚠ Alerta de stock bajo - {producto.nombre}"
    text_content = (
        f"El stock del producto {producto.nombre} (SKU {producto.sku}) en "
        f"{nombre_lugar} es inferior a 10 unidades. Quedan {stock_actual}."
    )
    html_content = render_to_string("emails/alerta_stock.html", context)

    msg = EmailMultiAlternatives(
        subject=subject,
        body=text_content,
        from_email=settings.EMAIL_HOST_USER,
        to=settings.TICKETS_NOTIFY_EMAILS,
    )
    msg.attach_alternative(html_content, "text/html")
    msg.send(fail_silently=False)

# Si tu modelo TipoMovimiento tiene un campo NOT NULL 'direccion',
# definimos un default por código:
def _direccion_default_value():
    """
    Devuelve el valor por defecto correcto para TipoMovimiento.direccion
    según su tipo: 0 si es IntegerField, "NEUTRO" si es CharField.
    """
    field = TipoMovimiento._meta.get_field("direccion")
    if isinstance(field, djmodels.IntegerField):
        return 0   # neutro numérico
    return "NEUTRO"  # neutro texto

def _tm(codigo: str) -> TipoMovimiento:
    """
    Devuelve (o crea) un TipoMovimiento con 'direccion' y 'nombre' completos,
    evitando IntegrityError por NOT NULL y el error de tipos.
    """
    codigo_up = (codigo or "").upper()
    defaults = {
        "nombre": codigo_up.replace("_", " ").title(),
        "direccion": _direccion_default_value(),
    }
    obj, created = TipoMovimiento.objects.get_or_create(codigo=codigo_up, defaults=defaults)

    # Completar si faltan valores o están vacíos
    to_update = []
    if getattr(obj, "direccion", None) in (None, ""):
        obj.direccion = defaults["direccion"]
        to_update.append("direccion")
    if not getattr(obj, "nombre", None):
        obj.nombre = defaults["nombre"]
        to_update.append("nombre")
    if to_update:
        obj.save(update_fields=to_update)

    return obj

def _unidad_default() -> UnidadMedida | None:
    return (
        UnidadMedida.objects.filter(codigo="UN").first()
        or UnidadMedida.objects.order_by("id").first()
    )
# ==== /Helpers Kardex ====
from itertools import chain
# ==== /Helpers Kardex ====

@login_required
def dashboard(request):
    alertas_stock = Stock.objects.filter(cantidad_disponible__lt=10).count()
    productos_totales = Producto.objects.count()
    categorias_count = CategoriaProducto.objects.count()

    sucursales = Sucursal.objects.all()
    bodegas = Bodega.objects.all()

    # STOCK TOTAL POR SUCURSAL
    stock_por_sucursal_qs = (
        Stock.objects
        .filter(ubicacion_sucursal__isnull=False)
        .values('ubicacion_sucursal__sucursal__nombre')
        .annotate(total=Sum('cantidad_disponible'))
        .order_by('ubicacion_sucursal__sucursal__nombre')
    )

    # STOCK TOTAL POR BODEGA
    stock_por_bodega_qs = (
        Stock.objects
        .filter(ubicacion_bodega__isnull=False)
        .values('ubicacion_bodega__bodega__nombre')
        .annotate(total=Sum('cantidad_disponible'))
        .order_by('ubicacion_bodega__bodega__nombre')
    )

    stock_por_sucursal = list(stock_por_sucursal_qs)
    stock_por_bodega = list(stock_por_bodega_qs)

    # % relativo dentro de cada grupo (para ancho de la barra)
    suc_totales = [float(i["total"]) for i in stock_por_sucursal]
    bod_totales = [float(i["total"]) for i in stock_por_bodega]

    max_suc = max(suc_totales) if suc_totales else 0
    max_bod = max(bod_totales) if bod_totales else 0

    for item in stock_por_sucursal:
        item["pct"] = int((float(item["total"]) / max_suc) * 100) if max_suc > 0 else 0

    for item in stock_por_bodega:
        item["pct"] = int((float(item["total"]) / max_bod) * 100) if max_bod > 0 else 0

    context = {
        "sucursales_activas": sucursales.count(),
        "bodegas_totales": bodegas.count(),
        "alertas_stock": alertas_stock,
        "productos_totales": productos_totales,
        "categorias_count": categorias_count,
        "sucursales": sucursales,
        "stock_por_sucursal": stock_por_sucursal,
        "stock_por_bodega": stock_por_bodega,
    }
    return render(request, "core/dashboard.html", context)

@login_required
def products(request):
    q = (request.GET.get("q") or "").strip()

    queryset = (
        Producto.objects.select_related("marca", "categoria", "unidad_base", "tasa_impuesto")
        .annotate(
            proveedores_count=Count("usuarios_proveedor", distinct=True),
            lotes_count=Count("lotes", distinct=True),
            series_count=Count("series", distinct=True),
            total_stock=Coalesce(
                Sum("stocks__cantidad_disponible"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            ),
        )
        .order_by("nombre", "sku")
    )

    if q:
        terms = [t for t in q.replace("-", " ").split() if t]
        for t in terms:
            queryset = queryset.filter(
                Q(sku__icontains=t)
                | Q(nombre__icontains=t)
                | Q(marca__nombre__icontains=t)
                | Q(categoria__nombre__icontains=t)
            )

    paginator = Paginator(queryset, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "core/products.html",
        {
            "q": q,
            "productos": page_obj.object_list,
            "page_obj": page_obj,
            "is_paginated": page_obj.has_other_pages(),
        },
    )

@login_required
def category(request, slug):
    return render(request, "core/category.html", {"category_name": slug.replace("-", " ").title()})


@login_required
def product_add(request):
    """
    versión nueva: el form de producto YA NO trae campo ubicacion.
    así que solo creamos el producto y listo.
    si quieres stock inicial por ubicación, hazlo en otra vista.
    """
    if request.method == "POST":
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save()
            messages.success(request, "Producto creado.")
            return redirect("products")
    else:
        form = ProductoForm()

    return render(request, "core/product_add.html", {"form": form})



def _get_or_create_default_location(bodega):
    """
    Devuelve la ubicación 'SIN UBICACIÓN' de una bodega,
    creándola si no existe.

    Código: <COD_BODEGA>-000-000
    Ej: BOD-01-000-000
    """
    if bodega is None:
        raise ValueError("Se llamó a _get_or_create_default_location sin bodega.")

    # Prefijo: código de bodega o, si no tiene, el ID
    prefijo = (getattr(bodega, "codigo", "") or str(bodega.pk)).strip().upper()

    area_cod = "000"
    estante_cod = "000"
    codigo = f"{prefijo}-{area_cod}-{estante_cod}"

    # 1) crear/obtener SOLO con los campos seguros
    ubi, created = UbicacionBodega.objects.get_or_create(
        bodega=bodega,
        codigo=codigo,
    )

    # 2) Rellenar los demás campos SOLO si existen en el modelo
    updated_fields = []

    if hasattr(ubi, "area_codigo") and ubi.area_codigo != area_cod:
        ubi.area_codigo = area_cod
        updated_fields.append("area_codigo")

    if hasattr(ubi, "estante_codigo") and ubi.estante_codigo != estante_cod:
        ubi.estante_codigo = estante_cod
        updated_fields.append("estante_codigo")

    if hasattr(ubi, "area") and not (ubi.area or "").strip():
        ubi.area = "SIN UBICACIÓN"
        updated_fields.append("area")

    if hasattr(ubi, "activo") and not ubi.activo:
        ubi.activo = True
        updated_fields.append("activo")

    if updated_fields:
        ubi.save(update_fields=updated_fields)

    return ubi





@login_required
@transaction.atomic
def product_add_combined(request):
    if request.method == "POST":
        pform = ProductoForm(request.POST, include_stock=False)
        sform = StockInlineForm(request.POST)

        if pform.is_valid() and sform.is_valid():
            # Si marca vencimiento, exige fecha
            if pform.cleaned_data.get("tiene_vencimiento") and not sform.cleaned_data.get("fecha_vencimiento"):
                sform.add_error("fecha_vencimiento", "Debes indicar una fecha de vencimiento para este producto.")
            else:
                # 1) Crear el producto
                producto = pform.save()

                # 2) Stock inicial + bodega seleccionada
                bodega = sform.cleaned_data["bodega"]
                cantidad = sform.cleaned_data["cantidad_inicial"] or 0

                stock_field_names = {f.name for f in Stock._meta.get_fields()}
                stock_kwargs = dict(
                    producto=producto,
                    cantidad_disponible=cantidad,
                )

                if "bodega" in stock_field_names:
                    # Caso 1: FK directa a Bodega
                    stock_kwargs["bodega"] = bodega
                else:
                    # Caso 2: usar ubicacion_bodega por defecto en esa bodega
                    ubi = _get_or_create_default_location(bodega)
                    if ubi and "ubicacion_bodega" in stock_field_names:
                        stock_kwargs["ubicacion_bodega"] = ubi

                # Crear el stock para el producto en la bodega
                Stock.objects.create(**stock_kwargs)

                # 3) Asignar stock a todas las sucursales de la bodega
                sucursales = Sucursal.objects.filter(bodega=bodega)
                for sucursal in sucursales:
                    ubi_sucursal = ensure_ubicacion_sucursal(sucursal)
                    Stock.objects.get_or_create(
                        producto=producto,
                        ubicacion_sucursal=ubi_sucursal,
                        defaults={"cantidad_disponible": Decimal("0")}  # Inicializamos en 0 si no hay cantidad
                    )

                # (opcional) recalcula stock global
                try:
                    total = (Stock.objects
                             .filter(producto=producto)
                             .aggregate(total=Sum("cantidad_disponible"))["total"]) or 0
                    producto.stock = total
                    producto.save(update_fields=["stock"])
                except Exception:
                    pass

                messages.success(request, "Producto y stock inicial registrados correctamente.")
                return redirect(reverse("producto-detail", args=[producto.id]))
    else:
        pform = ProductoForm(include_stock=False)
        sform = StockInlineForm()

    return render(request, "core/product_add.html", {"pform": pform, "sform":sform})
                                                     



# -------------------- Login Helpers --------------------
def _redirect_url_by_role(perfil):
    if not perfil or not perfil.rol:
        return reverse('dashboard')
    mapping = {
        'ADMIN': reverse('dashboard'),
        'BODEGUERO': reverse('products'),
    }
    return mapping.get(perfil.rol, reverse('dashboard'))


# -------------------- Login / Logout --------------------
def login_view(request):
    # Si ya estÃ¡ logueado, redirige segÃºn su rol
    if request.user.is_authenticated:
        perfil = getattr(request.user, 'perfil', None)
        return redirect(_redirect_url_by_role(perfil))

    next_url = request.GET.get('next') or request.POST.get('next')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        remember = request.POST.get('remember') == 'on'

        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            login(request, user)

            # "Recordarme": si NO marca, expira al cerrar el navegador
            if not remember:
                request.session.set_expiry(0)

            perfil = getattr(user, 'perfil', None)
            return redirect(next_url or _redirect_url_by_role(perfil))
        else:
            return render(request, 'accounts/login.html', {
                'error': 'Usuario o contraseÃ±a incorrectos',
                'next': next_url,
            })

    return render(request, 'accounts/login.html', {'next': next_url})

def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def dashboard_view(request):
    perfil = getattr(request.user, 'perfil', None)
    return redirect(_redirect_url_by_role(perfil))



# -------------------- Signup (opcional, solo ADMIN) --------------------
class SignupUserForm(UserCreationForm):
    email = forms.EmailField(required=False, label="Email")
    first_name = forms.CharField(required=False, label="Nombre")
    last_name = forms.CharField(required=False, label="Apellido")

    class Meta:
        model = User
        fields = ("username", "first_name", "last_name", "email", "password1", "password2")


class UsuarioPerfilForm(forms.ModelForm):
    class Meta:
        model = UsuarioPerfil
        fields = ("telefono",)  # aÃ±ade mÃ¡s campos si quieres capturarlos al alta


def _is_admin(user: User) -> bool:
    try:
        if not user or not user.is_authenticated:
            return False
        # Superusuario del sistema siempre tiene permisos de ADMIN
        if getattr(user, 'is_superuser', False):
            return True
        # Asegura que exista perfil para usuarios creados de forma externa
        perfil, _ = UsuarioPerfil.objects.get_or_create(usuario=user)
        return perfil.rol == UsuarioPerfil.Rol.ADMIN
    except Exception:
        return False
    
def _sync_user_groups_by_profile(user: User):
    """
    Sincroniza el Group del usuario segÃºn su perfil.rol.
    """
    try:
        perfil = user.perfil
        if perfil and perfil.rol:
            # Asegura que existan todos los grupos
            for code, _ in UsuarioPerfil.Rol.choices:
                Group.objects.get_or_create(name=code)
            user.groups.clear()
            user.groups.add(Group.objects.get(name=perfil.rol))
    except Exception:
        pass


def admin_required(view_func):
    """Decorador para views basadas en funciones que redirige con mensaje si no es admin."""
    def _wrapped(request, *args, **kwargs):
        if not _is_admin(request.user):
            messages.error(request, "No tienes permisos para realizar esa acciÃ³n.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return _wrapped

@admin_required
@transaction.atomic
def signup(request):
    if request.method == "POST":
        user_form = SignupUserForm(request.POST)
        perfil_form = UsuarioPerfilForm(request.POST)

        if user_form.is_valid() and perfil_form.is_valid():
            user = user_form.save(commit=True)
            perfil, _ = UsuarioPerfil.objects.get_or_create(usuario=user)

            for field, value in perfil_form.cleaned_data.items():
                setattr(perfil, field, value)
            perfil.save()

            login(request, user)
            messages.success(request, "âœ… Usuario creado correctamente.")
            try:
                return redirect(reverse("usuario-list"))
            except NoReverseMatch:
                return redirect("dashboard")
        else:
            messages.error(request, "âŒ Revisa los errores del formulario.")
    else:
        user_form = SignupUserForm()
        perfil_form = UsuarioPerfilForm()

    return render(request, "accounts/sign.html", {"user_form": user_form, "perfil_form": perfil_form})
@admin_required
@transaction.atomic
def user_create(request):
    """
    Alta de usuario (User + UsuarioPerfil).
    Usa: accounts/sign.html (tu formulario existente).
    Si viene ?role=PROVEEDOR, se preselecciona ese rol.
    """
    preset_role = request.GET.get("role")
    if request.method == "POST":
        user_form = SignupUserForm(request.POST)
        perfil_form = UsuarioPerfilForm(request.POST)
        if user_form.is_valid() and perfil_form.is_valid():
            user = user_form.save(commit=True)

            perfil, _ = UsuarioPerfil.objects.get_or_create(usuario=user)
            # copiar campos del form al perfil
            for field, value in perfil_form.cleaned_data.items():
                setattr(perfil, field, value)

            # si vino role por query y no se cambiÃ³ en el form, respÃ©talo
            if preset_role and not perfil_form.cleaned_data.get("rol"):
                perfil.rol = preset_role

            perfil.save()
            _sync_user_groups_by_profile(user)

            messages.success(request, "âœ… Usuario creado correctamente.")
            return redirect("usuario-list")
        messages.error(request, "âŒ Revisa los errores del formulario.")
    else:
        user_form = SignupUserForm()
        # inicializa el rol si viene por query
        initial = {}
        if preset_role in dict(UsuarioPerfil.Rol.choices):
            initial["rol"] = preset_role
        perfil_form = UsuarioPerfilForm(initial=initial)

    return render(request, "accounts/sign.html", {
        "user_form": user_form,
        "perfil_form": perfil_form,
    })

@admin_required
@transaction.atomic
def user_edit(request, user_id: int):
    obj = get_object_or_404(User.objects.select_related("perfil"), pk=user_id)

    # Evitar que un admin se desactive/elimine a sÃ­ mismo por accidente (opcional)
    editing_self = (request.user.pk == obj.pk)

    # Asegurar que tenga perfil
    perfil, _ = UsuarioPerfil.objects.get_or_create(usuario=obj)

    if request.method == "POST":
        uform = UserEditForm(request.POST, instance=obj)
        pform = UsuarioPerfilEditForm(request.POST, instance=perfil)

        if uform.is_valid() and pform.is_valid():
            # No permitir que un usuario se desactive a sÃ­ mismo (opcional, seguridad)
            if editing_self and not uform.cleaned_data.get("is_active", True):
                messages.error(request, "No puedes desactivar tu propio usuario.")
            else:
                uform.save()
                pform.save()
                _sync_user_groups_by_profile(obj)
                messages.success(request, "âœ… Usuario actualizado.")
                return redirect("usuario-list")
        else:
            messages.error(request, "âŒ Revisa los errores del formulario.")
    else:
        uform = UserEditForm(instance=obj)
        pform = UsuarioPerfilEditForm(instance=perfil)

    return render(request, "accounts/user_form.html", {
        "obj": obj,
        "user_form": uform,
        "perfil_form": pform,
    })

@admin_required
@transaction.atomic
def user_delete(request, user_id: int):
    obj = get_object_or_404(User, pk=user_id)

    # Reglas de seguridad Ãºtiles:
    if request.user.pk == obj.pk:
        messages.error(request, "No puedes eliminar tu propio usuario.")
        return redirect("usuario-list")
    if obj.is_superuser:
        messages.error(request, "No puedes eliminar un superusuario.")
        return redirect("usuario-list")

    if request.method == "POST":
        obj.delete()
        messages.success(request, "ðŸ—‘ï¸ Usuario eliminado.")
        return redirect("usuario-list")

    return render(request, "accounts/user_confirm_delete.html", {"obj": obj})

@admin_required
@login_required
def user_list(request):
    """
    Listado con bÃºsqueda y filtro por rol.
    Template: accounts/user_list.html
    """
    q = (request.GET.get("q") or "").strip()
    rol = (request.GET.get("rol") or "").strip()

    qs = User.objects.select_related("perfil").order_by("username")

    if q:
        qs = qs.filter(
            Q(username__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(email__icontains=q)
        )

    if rol:
        # rol es un CharField con choices en UsuarioPerfil
        qs = qs.filter(perfil__rol=rol)

    paginator = Paginator(qs, 20)
    page = request.GET.get("page", 1)
    users_page = paginator.get_page(page)

    return render(
        request,
        "accounts/user_list.html",
        {
            "users": users_page,
            "q": q,
            "rol": rol,
            "roles": UsuarioPerfil.Rol.choices,  # (code, label)
        },
    )


@admin_required
@login_required
@require_POST
def usuario_set_rol(request, user_id: int):
    """
    Cambia el rol (perfil.rol) de un usuario vÃ­a fetch POST JSON.
    Espera: {"rol": "<CODE>"}   (donde CODE es uno de UsuarioPerfil.Rol.choices)
    Responde: {"ok": true, "rol_label": "<Etiqueta>"}  o {"ok": false, "error": "..."}
    """
    import json

    try:
        payload = json.loads(request.body or "{}")
        code = (payload.get("rol") or "").strip()
        if not code:
            return JsonResponse({"ok": False, "error": "Rol no especificado."}, status=400)

        # Validar que el code estÃ© en choices
        valid_map = dict(UsuarioPerfil.Rol.choices)   # {code: label}
        if code not in valid_map:
            return JsonResponse({"ok": False, "error": "CÃ³digo de rol invÃ¡lido."}, status=400)

        target = get_object_or_404(User, pk=user_id)

        # Protecciones: no tocar superuser ni al propio usuario
        if target.is_superuser or target.id == request.user.id:
            return JsonResponse({"ok": False, "error": "No puedes cambiar este rol."}, status=403)

        perfil = getattr(target, "perfil", None)
        if perfil is None:
            return JsonResponse({"ok": False, "error": "El usuario no tiene perfil."}, status=400)

        # Guardar el CharField
        perfil.rol = code
        perfil.save(update_fields=["rol"])

        return JsonResponse({"ok": True, "rol_label": valid_map[code]})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)

# -------------------- CRUD Sucursal / Bodega --------------------
class AdminOnlyMixin(UserPassesTestMixin):
    """Mixin para CBV que deja pasar sÃ³lo a ADMIN."""
    def test_func(self):
        return _is_admin(self.request.user)
    
    def handle_no_permission(self):
        # Redirige a dashboard con mensaje en lugar de mostrar 403
        messages.error(self.request, "No tienes permisos para realizar esa acciÃ³n.")
        return redirect('dashboard')


class BodegaPermissionMixin(UserPassesTestMixin):
    """Permite acceso si es ADMIN o BODEGUERO (validaciÃ³n por objeto aparte)."""
    def test_func(self):
        try:
            perfil = getattr(self.request.user, 'perfil', None)
            if not self.request.user.is_authenticated:
                return False
            if perfil and perfil.rol == UsuarioPerfil.Rol.ADMIN:
                return True
            if perfil and perfil.rol == UsuarioPerfil.Rol.BODEGUERO:
                return True
            return False
        except Exception:
            return False

    def handle_no_permission(self):
        messages.error(self.request, "No tienes permisos para realizar esa acciÃ³n.")
        return redirect('dashboard')


def admin_required(view_func):
    """Decorador para views basadas en funciones que redirige con mensaje si no es admin."""
    def _wrapped(request, *args, **kwargs):
        if not _is_admin(request.user):
            messages.error(request, "No tienes permisos para realizar esa acciÃ³n.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return _wrapped

class SucursalListView(LoginRequiredMixin, ListView):
    model = Sucursal
    template_name = "core/sucursal_list.html"
    context_object_name = "sucursales"
    paginate_by = 20  # por defecto

    def get_paginate_by(self, queryset):
        """
        Permite ?page_size= en la URL (mÃ¡x 100).
        """
        try:
            size = int(self.request.GET.get("page_size", self.paginate_by))
        except (TypeError, ValueError):
            size = self.paginate_by
        return max(1, min(size, 100))

    def get_queryset(self):
        q = (self.request.GET.get("q") or "").strip()

        qs = (
            Sucursal.objects
            .select_related("bodega")
            .prefetch_related(
                "ubicaciones",                 # ubicaciones de la sucursal
                "ubicaciones__stocks__producto",  # stocks + producto
            )
            .annotate(
                total_stock=Coalesce(
                    Sum("ubicaciones__stocks__cantidad_disponible"),
                    Value(0, output_field=DecimalField(max_digits=20, decimal_places=6))
                ),
                ubicaciones_count=Count("ubicaciones", distinct=True),
                # cuenta productos distintos con stock > 0 en esa sucursal
                productos_count=Count(
                    "ubicaciones__stocks__producto",
                    filter=Q(ubicaciones__stocks__cantidad_disponible__gt=0),
                    distinct=True,
                ),
            )
            .only(
                "id", "codigo", "nombre", "ciudad", "region", "pais",
                "activo", "bodega__codigo", "bodega__nombre"
            )
            .order_by(Lower("codigo").asc())
        )

        if q:
            qs = qs.filter(
                Q(codigo__icontains=q) |
                Q(nombre__icontains=q) |
                Q(ciudad__icontains=q) |
                Q(region__icontains=q) |
                Q(bodega__codigo__icontains=q) |
                Q(bodega__nombre__icontains=q)
            )

        return qs

        def get_context_data(self, **kwargs):
            ctx = super().get_context_data(**kwargs)
            q = (self.request.GET.get("q") or "").strip()
            ctx["q"] = q
            ctx["has_filters"] = bool(q)
            ctx["total"] = self.get_queryset().count()
            ctx["page_size"] = self.get_paginate_by(self.get_queryset())
            return ctx

class SucursalCreateView(LoginRequiredMixin, AdminOnlyMixin, SuccessMessageMixin, CreateView):
    model = Sucursal
    form_class = SucursalForm
    template_name = "core/sucursal_form.html"
    success_message = "Sucursal creada correctamente."

    def get_success_url(self):
        return reverse_lazy("sucursal-list")


class SucursalUpdateView(LoginRequiredMixin, AdminOnlyMixin, SuccessMessageMixin, UpdateView):
    model = Sucursal
    form_class = SucursalForm
    template_name = "core/sucursal_form.html"
    success_message = "Sucursal actualizada correctamente."

    def get_success_url(self):
        return reverse_lazy("sucursal-list")


class SucursalDeleteView(LoginRequiredMixin, AdminOnlyMixin, SuccessMessageMixin, DeleteView):
    model = Sucursal
    template_name = "core/sucursal_confirm_delete.html"
    success_url = reverse_lazy("sucursal-list")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Sucursal eliminada correctamente.")
        return super().delete(request, *args, **kwargs)

class BodegaListView(LoginRequiredMixin, ListView):
    model = Bodega
    template_name = "core/bodega_list.html"
    context_object_name = "bodegas"
    paginate_by = 20

    def get_paginate_by(self, queryset):
        try:
            size = int(self.request.GET.get("page_size", self.paginate_by))
        except (TypeError, ValueError):
            size = self.paginate_by
        return max(1, min(size, 100))

    def get_queryset(self):
        q = (self.request.GET.get("q") or "").strip()

        qs = (
            Bodega.objects
            .prefetch_related("ubicaciones", "sucursales")
            .order_by(Lower("codigo").asc())
        )

        if q:
            qs = qs.filter(
                Q(codigo__icontains=q) |
                Q(nombre__icontains=q) |
                Q(descripcion__icontains=q) |
                Q(sucursales__nombre__icontains=q)
            ).distinct()

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        page_obj = ctx["page_obj"]
        bodegas_page = list(page_obj.object_list)

        # todas las ubicaciones de las bodegas de esta página
        ubi_ids = []
        for b in bodegas_page:
            for u in b.ubicaciones.all():
                ubi_ids.append(u.id)

        # traemos el stock de esas ubicaciones
        stock_qs = (
            Stock.objects
            .filter(ubicacion_bodega_id__in=ubi_ids)
            .select_related("producto", "ubicacion_bodega")
        )

        # indexar por id de ubicacion
        stock_por_ubi = {}
        for s in stock_qs:
            stock_por_ubi.setdefault(s.ubicacion_bodega_id, []).append(s)

        # inyectar a cada bodega
        for b in bodegas_page:
            detalle = []
            for u in b.ubicaciones.all():
                detalle.extend(stock_por_ubi.get(u.id, []))

            b.productos_con_stock = detalle
            b.total_stock = sum(d.cantidad_disponible for d in detalle)
            b.sucursales_count = b.sucursales.count()
            b.ubicaciones_count = b.ubicaciones.count()
            b.productos_count = len({d.producto_id for d in detalle})

        # para el <select> del modal
        ctx["productos"] = (
            Producto.objects
            .only("id", "sku", "nombre")
            .order_by("nombre")
        )

        ctx["q"] = (self.request.GET.get("q") or "").strip()
        return ctx

class BodegaCreateView(LoginRequiredMixin, BodegaPermissionMixin, SuccessMessageMixin, CreateView):
    model = Bodega
    form_class = BodegaForm
    template_name = "core/bodega_form.html"
    success_message = "Bodega creada correctamente."

    def get_success_url(self):
        return reverse_lazy("bodega-list")

    def get_form(self, form_class=None):
        """Restringe 'sucursal' para BODEGUERO (no para ADMIN/superuser)."""
        form = super().get_form(form_class)
        perfil = getattr(self.request.user, "perfil", None)

        # Superusuario o ADMIN: sin restricciÃ³n
        if self.request.user.is_superuser or (perfil and perfil.rol == UsuarioPerfil.Rol.ADMIN):
            return form

        # BODEGUERO: limitar queryset a su sucursal (o ninguno si no tiene)
        if perfil and perfil.rol == UsuarioPerfil.Rol.BODEGUERO:
            if getattr(perfil, "sucursal", None):
                form.fields["sucursal"].queryset = Sucursal.objects.filter(pk=perfil.sucursal.pk)
            else:
                form.fields["sucursal"].queryset = Sucursal.objects.none()
        return form

    def form_valid(self, form):
        """Fuerza la sucursal del BODEGUERO en el servidor (anti-manipulaciÃ³n)."""
        perfil = getattr(self.request.user, "perfil", None)
        if not self.request.user.is_superuser and perfil and perfil.rol == UsuarioPerfil.Rol.BODEGUERO:
            if getattr(perfil, "sucursal", None):
                form.instance.sucursal = perfil.sucursal
            else:
                messages.error(self.request, "No tienes una sucursal asignada.")
                return super().form_invalid(form)
        return super().form_valid(form)


class BodegaUpdateView(LoginRequiredMixin, BodegaPermissionMixin, SuccessMessageMixin, UpdateView):
    model = Bodega
    form_class = BodegaForm
    template_name = "core/bodega_form.html"
    success_message = "Bodega actualizada correctamente."

    def get_success_url(self):
        return reverse_lazy("bodega-list")

    def dispatch(self, request, *args, **kwargs):
        """BODEGUERO solo puede editar bodegas (salvo superuser)."""
        perfil = getattr(request.user, "perfil", None)
        if not request.user.is_superuser and perfil and perfil.rol == UsuarioPerfil.Rol.BODEGUERO:
            obj = self.get_object()
            if not obj or obj.sucursal_id != (perfil.sucursal.id if getattr(perfil, "sucursal", None) else None):
                messages.error(request, "No tienes permisos para editar esta bodega.")
                return redirect("bodega-list")
        return super().dispatch(request, *args, **kwargs)

    def get_form(self, form_class=None):
        """Restringe select de 'sucursal' para BODEGUERO; ADMIN/superuser ven todo."""
        form = super().get_form(form_class)
        perfil = getattr(self.request.user, "perfil", None)

        if self.request.user.is_superuser or (perfil and perfil.rol == UsuarioPerfil.Rol.ADMIN):
            return form

        if perfil and perfil.rol == UsuarioPerfil.Rol.BODEGUERO:
            if getattr(perfil, "sucursal", None):
                form.fields["sucursal"].queryset = Sucursal.objects.filter(pk=perfil.sucursal.pk)
            else:
                form.fields["sucursal"].queryset = Sucursal.objects.none()
        return form

class BodegaDeleteView(LoginRequiredMixin, BodegaPermissionMixin, SuccessMessageMixin, DeleteView):
    model = Bodega
    template_name = "core/bodega_confirm_delete.html"
    success_url = reverse_lazy("bodega-list")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Bodega eliminada correctamente.")
        return super().delete(request, *args, **kwargs)

    def dispatch(self, request, *args, **kwargs):
        """BODEGUERO solo puede eliminar bodegas de su sucursal (salvo superuser)."""
        perfil = getattr(request.user, "perfil", None)
        if not request.user.is_superuser and perfil and perfil.rol == UsuarioPerfil.Rol.BODEGUERO:
            obj = self.get_object()
            if not obj or obj.sucursal_id != (perfil.sucursal.id if getattr(perfil, "sucursal", None) else None):
                messages.error(request, "No tienes permisos para eliminar esta bodega.")
                return redirect("bodega-list")
        return super().dispatch(request, *args, **kwargs)


class BodegaDetailView(LoginRequiredMixin, DetailView):
    model = Bodega
    template_name = "core/bodega_detail.html"
    context_object_name = "bodega"

#Area de productos

from core.indicadores import get_eur_clp, get_utm_clp  # Importamos las funciones para obtener las tasas


class ProductsListView(LoginRequiredMixin, ListView):
    model = Producto
    template_name = "core/products.html"
    context_object_name = "productos"
    paginate_by = 20

    def get_queryset(self):
        q = (self.request.GET.get("q") or "").strip()

        qs = (
            Producto.objects
            .select_related("marca", "categoria", "unidad_base")
            .prefetch_related(
                "stocks_producto",
                "stocks_producto__ubicacion",
                "stocks_producto__ubicacion__bodega",
                "stocks_producto__ubicacion__sucursal",
            )
            .annotate(
                total_stock=Coalesce(
                    Sum("stocks_producto__cantidad_disponible"),
                    Value(0, output_field=DecimalField(max_digits=20, decimal_places=6))
                ),
                proveedores_count=Count("usuarios_proveedor", distinct=True),
                lotes_count=Count("lotes", distinct=True),
                series_count=Count("series", distinct=True),
            )
            .order_by("nombre")
        )

        if q:
            qs = qs.filter(
                Q(sku__icontains=q)
                | Q(nombre__icontains=q)
                | Q(marca__nombre__icontains=q)
                | Q(categoria__nombre__icontains=q)
            )

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = (self.request.GET.get("q") or "").strip()

        # Obtener las tasas de cambio de mindicador.cl
        eur = get_eur_clp()  # CLP por 1 EUR
        utm = get_utm_clp()  # CLP por 1 UTM

        # Calculamos el precio en EUR y UTM en la vista, no en el template
        try:
            ctx["productos"] = [
                {
                    "producto": producto,
                    "precio_eur": round(producto.precio / eur, 2) if eur else None,
                    "precio_utm": round(producto.precio / utm, 4) if utm else None,
                }
                for producto in self.get_queryset()
            ]
        except Exception as e:
            print(f"Error calculando los precios: {e}")
            # En caso de error dejamos los valores en None
            ctx["productos"] = self.get_queryset()

        return ctx






class ProductUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Producto
    form_class = ProductoForm
    template_name = "core/product_add.html"
    success_message = "Producto actualizado correctamente."

    def form_valid(self, form):
        resp = super().form_valid(form)
        nxt = self.request.POST.get("next") or self.request.GET.get("next")
        if nxt:
            return redirect(nxt)
        return resp

    def get_success_url(self):
        return reverse_lazy("products")

class ProductUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):

    model = Producto
    form_class = ProductoForm
    template_name = "core/product_update.html"
    context_object_name = "producto"
    success_message = "Producto actualizado correctamente."

    # ---------- Queryset “rápido” ----------
    def get_queryset(self):
        return (Producto.objects
                .select_related("marca", "categoria", "unidad_base", "tasa_impuesto"))

  
    pk_url_kwarg = "pk"
    slug_url_kwarg = "sku"
    slug_field = "sku"

    def get_object(self, queryset=None):
        qs = queryset or self.get_queryset()
        # 1) Prioridad: si hay <pk> en la URL, úsalo
        pk = self.kwargs.get(self.pk_url_kwarg)
        if pk is not None:
            try:
                return qs.get(pk=pk)
            except Producto.DoesNotExist:
                raise Http404("Producto no encontrado.")
        # 2) Si hay <sku> en la URL, úsalo
        sku_kw = self.kwargs.get(self.slug_url_kwarg)
        if sku_kw:
            try:
                return qs.get(sku=str(sku_kw).upper())
            except Producto.DoesNotExist:
                raise Http404("Producto no encontrado.")
        # 3) Fallback: si llega ?sku= en querystring
        sku_qs = (self.request.GET.get("sku") or "").strip().upper()
        if sku_qs:
            try:
                return qs.get(sku=sku_qs)
            except Producto.DoesNotExist:
                raise Http404("Producto no encontrado.")
        return super().get_object(qs)

    # ---------- Pasar flag al form ----------
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Editar solo metadatos del producto (no stock)
        kwargs["include_stock"] = False
        return kwargs

    # ---------- UX de errores ----------
    def form_invalid(self, form):
        # muestra un resumen arriba sin obligarte a inspeccionar campo por campo
        errors = []
        for name, field_errors in form.errors.items():
            for e in field_errors:
                errors.append(f"{name}: {e}")
        if errors:
            messages.error(self.request, "No se pudo actualizar el producto. Revisa los campos.")
        return super().form_invalid(form)

    # ---------- Redirecciones ----------
    def form_valid(self, form):
        resp = super().form_valid(form)
        nxt = self.request.POST.get("next") or self.request.GET.get("next")
        if nxt:
            return redirect(nxt)
        return resp

    def get_success_url(self):
        # Si hay ?next= úsalo, si no, vuelve al detalle del propio producto
        nxt = self.request.POST.get("next") or self.request.GET.get("next")
        if nxt:
            return nxt
        try:
            return reverse("producto-detail", args=[self.object.pk])
        except Exception:
            return reverse_lazy("products")



class ProductDeleteView(LoginRequiredMixin, DeleteView):
    model = Producto
    template_name = "core/product_confirm_delete.html"  # fallback si navegas directo
    success_url = reverse_lazy("products")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Producto eliminado correctamente.")
        return super().delete(request, *args, **kwargs)


from decimal import Decimal
from django.db.models import (
    Sum, Value, DecimalField, CharField, F, Case, When
)
from django.db.models.functions import Coalesce
# ...

class ProductDetailView(LoginRequiredMixin, DetailView):
    """
    Detalle de producto con:
      - info general del producto
      - métricas (stock total disponible, precio, estado)
      - desglose de stock por sucursal/bodega/ubicación
    """
    model = Producto
    template_name = "core/product_detail.html"
    context_object_name = "producto"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        producto = self.object

        # Totales (si no hay reservas, neto = disponible)
        agg = (
            Stock.objects
            .filter(producto=producto)
            .aggregate(
                total_disponible=Coalesce(
                    Sum("cantidad_disponible"),
                    Value(0, output_field=DecimalField(max_digits=20, decimal_places=6))
                )
            )
        )

        total_disponible = agg["total_disponible"] or Decimal("0")
        precio_unitario = producto.precio or Decimal("0")
        valor_total = precio_unitario * total_disponible

        ctx["totales"] = {
            "total_disponible": total_disponible,
            "total_neto": total_disponible,   # lo dejo por compatibilidad
            "precio_unitario": precio_unitario,
            "valor_total": valor_total,
        }

        # ==== resumen por sucursal/bodega (igual que ya tenías) ====
        resumen = (
            Stock.objects
            .filter(producto=producto)
            .annotate(
                sucursal_codigo=Case(
                    When(ubicacion_sucursal__isnull=False, then=F("ubicacion_sucursal__sucursal__codigo")),
                    default=Value("", output_field=CharField()),
                ),
                sucursal_nombre=Case(
                    When(ubicacion_sucursal__isnull=False, then=F("ubicacion_sucursal__sucursal__nombre")),
                    default=Value("", output_field=CharField()),
                ),
                bodega_codigo=Case(
                    When(ubicacion_bodega__isnull=False, then=F("ubicacion_bodega__bodega__codigo")),
                    default=Value("", output_field=CharField()),
                ),
                bodega_nombre=Case(
                    When(ubicacion_bodega__isnull=False, then=F("ubicacion_bodega__bodega__nombre")),
                    default=Value("", output_field=CharField()),
                ),
                sucursal_ubi_codigo=Case(
                    When(ubicacion_sucursal__isnull=False, then=F("ubicacion_sucursal__codigo")),
                    default=Value("", output_field=CharField()),
                ),
                sucursal_ubi_nombre=Case(
                    When(ubicacion_sucursal__isnull=False, then=F("ubicacion_sucursal__area")),
                    default=Value("", output_field=CharField()),
                ),
                bodega_ubi_codigo=Case(
                    When(ubicacion_bodega__isnull=False, then=F("ubicacion_bodega__codigo")),
                    default=Value("", output_field=CharField()),
                ),
                bodega_ubi_nombre=Case(
                    When(ubicacion_bodega__isnull=False, then=F("ubicacion_bodega__area")),
                    default=Value("", output_field=CharField()),
                ),
            )
            .values(
                "sucursal_codigo", "sucursal_nombre",
                "bodega_codigo", "bodega_nombre",
                "sucursal_ubi_codigo", "sucursal_ubi_nombre",
                "bodega_ubi_codigo", "bodega_ubi_nombre",
            )
            .annotate(
                total_disponible=Coalesce(
                    Sum("cantidad_disponible"),
                    Value(0, output_field=DecimalField(max_digits=20, decimal_places=6))
                ),
                total_neto=F("total_disponible"),
            )
            .order_by(
                "sucursal_codigo",
                "bodega_codigo",
                "sucursal_ubi_codigo",
                "bodega_ubi_codigo",
            )
        )

        ctx["resumen_sucursales"] = resumen
        return ctx


# ----------------------
# CRUD Ubicacion (pÃ¡ginas)
# ----------------------
class UbicacionBodegaListView(LoginRequiredMixin, ListView):
    model = UbicacionBodega
    template_name = "core/ubicacion_bodega_list.html"
    context_object_name = "ubicaciones"
    paginate_by = 20

    def get_queryset(self):
        q = (self.request.GET.get("q") or "").strip()
        bodega_id = (self.request.GET.get("bodega") or "").strip()

        qs = (
            UbicacionBodega.objects.select_related("bodega", "tipo")
            .annotate(
                total_stock=Coalesce(
                    Sum("stocks__cantidad_disponible"),
                    Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
                )
            )
            .order_by("bodega__codigo", "codigo")
        )

        if q:
            qs = qs.filter(Q(codigo__icontains=q) | Q(nombre__icontains=q) | Q(area__icontains=q))

        if bodega_id:
            qs = qs.filter(bodega_id=bodega_id)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["bodegas"] = Bodega.objects.all().order_by("codigo")
        ctx["q"] = (self.request.GET.get("q") or "").strip()
        ctx["f_bodega"] = (self.request.GET.get("bodega") or "").strip()
        return ctx


class UbicacionSucursalListView(LoginRequiredMixin, ListView):
    model = UbicacionSucursal
    template_name = "core/ubicacion_sucursal_list.html"
    context_object_name = "ubicaciones"
    paginate_by = 20

    def get_queryset(self):
        q = (self.request.GET.get("q") or "").strip()
        sucursal_id = (self.request.GET.get("sucursal") or "").strip()

        qs = (
            UbicacionSucursal.objects.select_related("sucursal", "tipo")
            .annotate(
                total_stock=Coalesce(
                    Sum("stocks__cantidad_disponible"),
                    Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
                )
            )
            .order_by("sucursal__codigo", "codigo")
        )

        if q:
            qs = qs.filter(Q(codigo__icontains=q) | Q(nombre__icontains=q) | Q(area__icontains=q))

        if sucursal_id:
            qs = qs.filter(sucursal_id=sucursal_id)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["sucursales"] = Sucursal.objects.all().order_by("codigo")
        ctx["q"] = (self.request.GET.get("q") or "").strip()
        ctx["f_sucursal"] = (self.request.GET.get("sucursal") or "").strip()
        return ctx


class UbicacionBodegaCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = UbicacionBodega
    form_class = UbicacionBodegaForm
    template_name = "core/ubicacion_form.html"
    success_message = "Ubicación de bodega creada."
    success_url = reverse_lazy("ubicacion-bodega-list")


class UbicacionBodegaUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = UbicacionBodega
    form_class = UbicacionBodegaForm
    template_name = "core/ubicacion_form.html"
    success_message = "Ubicación de bodega actualizada."
    success_url = reverse_lazy("ubicacion-bodega-list")


class UbicacionBodegaDeleteView(LoginRequiredMixin, DeleteView):
    model = UbicacionBodega
    template_name = "core/confirm_delete.html"
    success_url = reverse_lazy("ubicacion-bodega-list")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Ubicación eliminada.")
        return super().delete(request, *args, **kwargs)


class UbicacionSucursalCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = UbicacionSucursal
    form_class = UbicacionSucursalForm
    template_name = "core/ubicacion_form.html"
    success_message = "Ubicación de sucursal creada."
    success_url = reverse_lazy("ubicacion-sucursal-list")


class UbicacionSucursalUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = UbicacionSucursal
    form_class = UbicacionSucursalForm
    template_name = "core/ubicacion_form.html"
    success_message = "Ubicación de sucursal actualizada."
    success_url = reverse_lazy("ubicacion-sucursal-list")


class UbicacionSucursalDeleteView(LoginRequiredMixin, DeleteView):
    model = UbicacionSucursal
    template_name = "core/confirm_delete.html"
    success_url = reverse_lazy("ubicacion-sucursal-list")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Ubicación eliminada.")
        return super().delete(request, *args, **kwargs)

# ------------------------------------
#   TipoUbicacion con modal
# ------------------------------------
# Los modales usan <dialog> y cargan estas vistas que devuelven la pÃ¡gina completa,
# pero con templates chicos pensados para presentarse en un modal.

class TipoUbicacionCreateModal(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = TipoUbicacion
    form_class = TipoUbicacionForm
    template_name = "core/partials/tipo_form_modal.html"
    success_message = "Tipo de ubicaciÃ³n creado."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("bodega-list")


class TipoUbicacionUpdateModal(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = TipoUbicacion
    form_class = TipoUbicacionForm
    template_name = "core/partials/tipo_form_modal.html"
    success_message = "Tipo de ubicaciÃ³n actualizado."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("bodega-list")


class TipoUbicacionDeleteModal(LoginRequiredMixin, DeleteView):
    model = TipoUbicacion
    template_name = "core/partials/confirm_modal.html"
    success_url = reverse_lazy("bodega-list")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Tipo de ubicaciÃ³n eliminado.")
        return super().delete(request, *args, **kwargs)

# ======================
# Marca
# ======================
class MarcaCreateModal(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Marca
    form_class = MarcaForm
    template_name = "core/partials/marca_form_modal.html"
    success_message = "Marca creada."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("products")


class MarcaUpdateModal(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Marca
    form_class = MarcaForm
    template_name = "core/partials/marca_form_modal.html"
    success_message = "Marca actualizada."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("products")


class MarcaDeleteModal(LoginRequiredMixin, DeleteView):
    model = Marca
    template_name = "core/partials/confirm_modal.html"
    success_url = reverse_lazy("products")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Marca eliminada.")
        return super().delete(request, *args, **kwargs)

# ======================
# Unidad de Medida
# ======================
class UnidadMedidaCreateModal(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = UnidadMedida
    form_class = UnidadMedidaForm
    template_name = "core/partials/unidad_form_modal.html"
    success_message = "Unidad de medida creada."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("products")


class UnidadMedidaUpdateModal(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = UnidadMedida
    form_class = UnidadMedidaForm
    template_name = "core/partials/unidad_form_modal.html"
    success_message = "Unidad de medida actualizada."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("products")


class UnidadMedidaDeleteModal(LoginRequiredMixin, DeleteView):
    model = UnidadMedida
    template_name = "core/partials/confirm_modal.html"
    success_url = reverse_lazy("products")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Unidad de medida eliminada.")
        return super().delete(request, *args, **kwargs)

# ======================
# Tasa de Impuesto
# ======================
class TasaImpuestoCreateModal(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = TasaImpuesto
    form_class = TasaImpuestoForm
    template_name = "core/partials/tasa_form_modal.html"
    success_message = "Tasa de impuesto creada."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("products")


class TasaImpuestoUpdateModal(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = TasaImpuesto
    form_class = TasaImpuestoForm
    template_name = "core/partials/tasa_form_modal.html"
    success_message = "Tasa de impuesto actualizada."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("products")


class TasaImpuestoDeleteModal(LoginRequiredMixin, DeleteView):
    model = TasaImpuesto
    template_name = "core/partials/confirm_modal.html"
    success_url = reverse_lazy("products")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Tasa de impuesto eliminada.")
        return super().delete(request, *args, **kwargs)

# ======================
# CategorÃ­a de Producto
# ======================
class CategoriaProductoCreateModal(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = CategoriaProducto
    form_class = CategoriaProductoForm
    template_name = "core/partials/categoria_form_modal.html"
    success_message = "CategorÃ­a creada."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("products")


class CategoriaProductoUpdateModal(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = CategoriaProducto
    form_class = CategoriaProductoForm
    template_name = "core/partials/categoria_form_modal.html"
    success_message = "CategorÃ­a actualizada."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("products")


class CategoriaProductoDeleteModal(LoginRequiredMixin, DeleteView):
    model = CategoriaProducto
    template_name = "core/partials/confirm_modal.html"
    success_url = reverse_lazy("products")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "CategorÃ­a eliminada.")
        return super().delete(request, *args, **kwargs)
    
# ===== LoteProducto (modales) =====
class LoteCreateModal(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = LoteProducto
    form_class = LoteProductoForm
    template_name = "core/partials/lote_form_modal.html"
    success_message = "Lote creado."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("products")

class LoteUpdateModal(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = LoteProducto
    form_class = LoteProductoForm
    template_name = "core/partials/lote_form_modal.html"
    success_message = "Lote actualizado."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("products")

class LoteDeleteModal(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = LoteProducto
    template_name = "core/partials/confirm_delete_modal.html"
    success_message = "Lote eliminado."
    # messages en DeleteView requieren manejo en form_valid o post_delete signal; si usas messages en template, puedes mostrar el texto allÃ­.

# ===== SerieProducto (modales) =====
class SerieCreateModal(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = SerieProducto
    form_class = SerieProductoForm
    template_name = "core/partials/serie_form_modal.html"
    success_message = "Serie creada."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("products")

class SerieUpdateModal(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = SerieProducto
    form_class = SerieProductoForm
    template_name = "core/partials/serie_form_modal.html"
    success_message = "Serie actualizada."

    def get_success_url(self):
        return self.request.GET.get("next") or reverse_lazy("products")

class SerieDeleteModal(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = SerieProducto
    template_name = "core/partials/confirm_delete_modal.html"
    success_message = "Serie eliminada."

def test_scanner(request):
    return render(request, "core/test_scanner.html")


@login_required
def products(request):
    productos = (
        Producto.objects
        .select_related("marca", "categoria", "unidad_base", "tasa_impuesto")
        .annotate(
            lotes_count=Count("lotes", distinct=True),
            series_count=Count("series", distinct=True),
            proveedores_count=Count("productousuarioproveedor", distinct=True),
        )
        .order_by("nombre")
    )
    return render(request, "core/products.html", {
        "productos": productos,
        "q": (request.GET.get("q") or "").strip(),
    })

@login_required
def productos_por_bodega(request, bodega_id):
    try:
        bodega = Bodega.objects.get(id=bodega_id)
    except Bodega.DoesNotExist:
        return render(request, "core/error.html", {"error": "Bodega no encontrada."})

    # ubicaciones de esa bodega
    ubicaciones_en_bodega = bodega.ubicaciones.all()

    if not ubicaciones_en_bodega:
        return render(
            request,
            "core/error.html",
            {"error": "No se encontraron ubicaciones asociadas a esta bodega."},
        )

    # productos que tienen stock en alguna de esas ubicaciones
    productos = (
        Producto.objects.filter(stocks__ubicacion_bodega__bodega=bodega)
        .distinct()
        .annotate(
            stock_total=Coalesce(
                Sum("stocks__cantidad_disponible"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            )
        )
    )

    return render(
        request,
        "core/productos_por_bodega.html",
        {
            "bodega": bodega,
            "ubicaciones_en_bodega": ubicaciones_en_bodega,
            "productos": productos,
        },
    )


@login_required
def product_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = (
        Producto.objects
        .all()
        .annotate(
            # usa los nombres REALES que tienes en el modelo
            proveedores_count=Count("usuarios_proveedor", distinct=True),
            lotes_count=Count("lotes", distinct=True),
            series_count=Count("series", distinct=True),
        )
        .order_by("sku")
    )

    if q:
        qs = qs.filter(
            Q(sku__icontains=q) |
            Q(nombre__icontains=q) |
            Q(marca__nombre__icontains=q) |
            Q(categoria__nombre__icontains=q)
        )

    paginator = Paginator(qs, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "core/products.html",     # tu template
        {
            "q": q,
            "productos": page_obj.object_list,
            "page_obj": page_obj,
            "is_paginated": page_obj.has_other_pages(),
        },
    )

# ---------- Utilidades comunes ----------

def _is_fetch(request):
    # Tu helper JS manda 'X-Requested-With: fetch'
    return request.headers.get("X-Requested-With") == "fetch"

def _json_or_redirect(request, ok: bool, msg: str, redirect_to: str = None, extra: dict | None = None):
    payload = {"ok": ok, "message": msg}
    if extra:
        payload.update(extra)
    if _is_fetch(request):
        status = 200 if ok else 400
        return JsonResponse(payload, status=status)
    if ok:
        messages.success(request, msg)
    else:
        messages.error(request, msg)
    return redirect(redirect_to or request.META.get("HTTP_REFERER", "/"))
from decimal import Decimal

# === Config: pon aquÃ­ el nombre real del campo de reserva en Producto (si existe)
RESERVA_FIELD = "reserva"   # ej. "reservado" si tu modelo lo llama asÃ­


def _get_reserva(producto):
    """Obtiene la reserva desde Producto (0 si no existe el campo)."""
    return Decimal(getattr(producto, RESERVA_FIELD, 0) or 0)


def _set_reserva(producto, value):
    """Setea la reserva en Producto (si el campo existe). Ignora si no existe."""
    if hasattr(producto, RESERVA_FIELD):
        setattr(producto, RESERVA_FIELD, Decimal(value or 0))


def _recalcular_total_producto(producto):
    """
    VersiÃ³n sin tabla Stock.
    Asegura que 'stock' y 'reserva' existan y sean decimales >= 0.
    Ãštil como 'normalizador' tras operaciones.
    """
    disponible = Decimal(getattr(producto, "stock", 0) or 0)
    reservado  = _get_reserva(producto)

    if disponible < 0:
        disponible = Decimal(0)
    if reservado < 0:
        reservado = Decimal(0)

    producto.stock = disponible
    _set_reserva(producto, reservado)

    # Guardar solo los campos que existan
    update_fields = ["stock"]
    if hasattr(producto, RESERVA_FIELD):
        update_fields.append(RESERVA_FIELD)
    producto.save(update_fields=update_fields)


class _ProductoStockProxy:
    """
    Proxy de compatibilidad para cÃ³digo legacy que esperaba un objeto Stock.
    Lee/escribe directamente en Producto.stock y Producto.<RESERVA_FIELD>.
    Tiene un mÃ©todo .save(update_fields=...) para imitar la API.
    """
    def __init__(self, producto):
        self._p = producto

    @property
    def producto(self):
        return self._p

    @property
    def cantidad_disponible(self):
        return Decimal(self._p.stock or 0)

    @cantidad_disponible.setter
    def cantidad_disponible(self, value):
        self._p.stock = Decimal(value or 0)

    @property
    def cantidad_reservada(self):
        return _get_reserva(self._p)

    @cantidad_reservada.setter
    def cantidad_reservada(self, value):
        _set_reserva(self._p, value)

    def save(self, update_fields=None):
        # Mapea update_fields del "Stock" a campos reales de Producto
        fields = set(update_fields or [])
        mapped = set()
        if not fields:
            # Guardar ambos si existen
            mapped.add("stock")
            if hasattr(self._p, RESERVA_FIELD):
                mapped.add(RESERVA_FIELD)
        else:
            if "cantidad_disponible" in fields or "stock" in fields:
                mapped.add("stock")
            if "cantidad_reservada" in fields or RESERVA_FIELD in fields:
                if hasattr(self._p, RESERVA_FIELD):
                    mapped.add(RESERVA_FIELD)

        if not mapped:
            # fallback por seguridad
            mapped.add("stock")
            if hasattr(self._p, RESERVA_FIELD):
                mapped.add(RESERVA_FIELD)

        self._p.save(update_fields=list(mapped))


def _get_or_create_stock(producto, ubicacion_id: int | None = None):
    """
    Compatibilidad sin tabla Stock:
    Ignora la ubicaciÃ³n y devuelve un proxy ligado al Producto.
    AsÃ­ tu cÃ³digo legacy (que hacÃ­a .cantidad_disponible += x; .save()) sigue funcionando.
    """
    return _ProductoStockProxy(producto)


# ====== Helpers â€œnuevosâ€ para operar stock directamente en Producto ======

def ajustar_stock(producto, delta_disponible=0, delta_reservado=0, guardar=True):
    """
    Suma/resta cantidades directamente en Producto.
    Uso:
      ajustar_stock(prod, delta_disponible=+5)       # entrada de stock
      ajustar_stock(prod, delta_disponible=-2)       # salida de stock
      ajustar_stock(prod, delta_reservado=+3)        # reservar 3
      ajustar_stock(prod, delta_reservado=-1)        # liberar 1 de reserva
    """
    disponible = Decimal(getattr(producto, "stock", 0) or 0) + Decimal(delta_disponible or 0)
    reservado  = _get_reserva(producto) + Decimal(delta_reservado or 0)

    if disponible < 0:
        disponible = Decimal(0)
    if reservado < 0:
        reservado = Decimal(0)

    producto.stock = disponible
    _set_reserva(producto, reservado)

    if guardar:
        update_fields = ["stock"]
        if hasattr(producto, RESERVA_FIELD):
            update_fields.append(RESERVA_FIELD)
        producto.save(update_fields=update_fields)
    return producto


def set_stock(producto, disponible=None, reservado=None, guardar=True):
    """
    Seteo directo (absoluto) de stock/reserva en Producto.
    """
    if disponible is not None:
        d = Decimal(disponible or 0)
        producto.stock = d if d > 0 else Decimal(0)

    if reservado is not None and hasattr(producto, RESERVA_FIELD):
        r = Decimal(reservado or 0)
        _set_reserva(producto, r if r > 0 else Decimal(0))

    if guardar:
        update_fields = ["stock"]
        if hasattr(producto, RESERVA_FIELD):
            update_fields.append(RESERVA_FIELD)
        producto.save(update_fields=update_fields)
    return producto

# ---------- Consulta: ver stock por producto (tu lÃ³gica, con pequeÃ±os ajustes) ----------
from django.db.models import Sum, F



from django.db.models import (
    Sum, Value, DecimalField, F, Case, When, CharField
)
from django.db.models.functions import Coalesce



@login_required
def stock_por_producto(request):
    """
    Consulta de stock por SKU.
    Muestra:
      - totales globales
      - desglose por sucursal/bodega (resumen_sucursales)
      - detalle por ubicación interna (ubicaciones_detalle)
      - datos para gráfico de barras por ubicación
    """
    sku = (request.GET.get("sku") or "").strip().upper()
    producto = None
    totales = None
    resumen_sucursales = []
    ubicaciones_detalle = []
    chart_labels = []
    chart_values = []

    if sku:
        try:
            producto = (
                Producto.objects
                .select_related("marca", "categoria")
                .get(sku=sku)
            )

            # 1) TOTAL GLOBAL DEL PRODUCTO
            agg = (
                Stock.objects
                .filter(producto=producto)
                .aggregate(
                    total_disponible=Coalesce(
                        Sum("cantidad_disponible"),
                        Value(0, output_field=DecimalField(max_digits=20, decimal_places=6))
                    )
                )
            )
            disponible = agg["total_disponible"] or 0

            # Precio unitario y valor total (mismas claves que en product_detail)
            precio_unitario = producto.precio or Decimal("0")
            # nos aseguramos que disponible sea Decimal
            disponible_dec = Decimal(str(disponible))
            valor_total = precio_unitario * disponible_dec

            totales = {
                "total_disponible": disponible_dec,
                "total_neto": disponible_dec,       # por si algo aún lo usa
                "precio_unitario": precio_unitario,
                "valor_total": valor_total,
            }

            # 2) DESGLOSE POR SUCURSAL / BODEGA
            resumen_sucursales = (
                Stock.objects
                .filter(producto=producto)
                .annotate(
                    sucursal_codigo=Case(
                        When(
                            ubicacion_sucursal__isnull=False,
                            then=F("ubicacion_sucursal__sucursal__codigo"),
                        ),
                        default=Value("", output_field=CharField()),
                    ),
                    sucursal_nombre=Case(
                        When(
                            ubicacion_sucursal__isnull=False,
                            then=F("ubicacion_sucursal__sucursal__nombre"),
                        ),
                        default=Value("", output_field=CharField()),
                    ),
                    bodega_codigo=Case(
                        When(
                            ubicacion_bodega__isnull=False,
                            then=F("ubicacion_bodega__bodega__codigo"),
                        ),
                        default=Value("", output_field=CharField()),
                    ),
                    bodega_nombre=Case(
                        When(
                            ubicacion_bodega__isnull=False,
                            then=F("ubicacion_bodega__bodega__nombre"),
                        ),
                        default=Value("", output_field=CharField()),
                    ),
                )
                .values(
                    "sucursal_codigo",
                    "sucursal_nombre",
                    "bodega_codigo",
                    "bodega_nombre",
                )
                .annotate(
                    total_disponible=Coalesce(
                        Sum("cantidad_disponible"),
                        Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
                    ),
                    total_neto=F("total_disponible"),
                )
                .order_by("sucursal_codigo", "bodega_codigo")
            )

            # 3) DETALLE POR UBICACIÓN INTERNA (para tabla + gráfico)
            ubicaciones_qs = (
                Stock.objects
                .filter(producto=producto)
                .annotate(
                    sucursal_codigo=Case(
                        When(
                            ubicacion_sucursal__isnull=False,
                            then=F("ubicacion_sucursal__sucursal__codigo"),
                        ),
                        default=Value("", output_field=CharField()),
                    ),
                    sucursal_nombre=Case(
                        When(
                            ubicacion_sucursal__isnull=False,
                            then=F("ubicacion_sucursal__sucursal__nombre"),
                        ),
                        default=Value("", output_field=CharField()),
                    ),
                    bodega_codigo=Case(
                        When(
                            ubicacion_bodega__isnull=False,
                            then=F("ubicacion_bodega__bodega__codigo"),
                        ),
                        default=Value("", output_field=CharField()),
                    ),
                    bodega_nombre=Case(
                        When(
                            ubicacion_bodega__isnull=False,
                            then=F("ubicacion_bodega__bodega__nombre"),
                        ),
                        default=Value("", output_field=CharField()),
                    ),
                    ubicacion_codigo=Case(
                        When(
                            ubicacion_sucursal__isnull=False,
                            then=F("ubicacion_sucursal__codigo"),
                        ),
                        When(
                            ubicacion_bodega__isnull=False,
                            then=F("ubicacion_bodega__codigo"),
                        ),
                        default=Value("", output_field=CharField()),
                    ),
                    ubicacion_area=Case(
                        When(
                            ubicacion_sucursal__isnull=False,
                            then=F("ubicacion_sucursal__area"),
                        ),
                        When(
                            ubicacion_bodega__isnull=False,
                            then=F("ubicacion_bodega__area"),
                        ),
                        default=Value("", output_field=CharField()),
                    ),
                )
                .values(
                    "sucursal_codigo",
                    "sucursal_nombre",
                    "bodega_codigo",
                    "bodega_nombre",
                    "ubicacion_codigo",
                    "ubicacion_area",
                )
                .annotate(
                    total_disponible=Coalesce(
                        Sum("cantidad_disponible"),
                        Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
                    )
                )
                .order_by(
                    "sucursal_codigo",
                    "bodega_codigo",
                    "ubicacion_codigo",
                )
            )

            ubicaciones_detalle = list(ubicaciones_qs)

            # 4) Datos para el gráfico (labels + valores)
            for row in ubicaciones_detalle:
                label = row.get("ubicacion_codigo") or ""
                area = row.get("ubicacion_area") or ""
                if area:
                    label = f"{label} · {area}"
                chart_labels.append(label)
                chart_values.append(float(row.get("total_disponible") or 0))

        except Producto.DoesNotExist:
            messages.error(request, f"No se encontró ningún producto con SKU '{sku}'.")

    return render(
        request,
        "core/stock_producto.html",
        {
            "sku": sku,
            "producto": producto,
            "totales": totales,
            "resumen_sucursales": resumen_sucursales,
            "ubicaciones_detalle": ubicaciones_detalle,
            "chart_labels_json": json.dumps(chart_labels),
            "chart_values_json": json.dumps(chart_values),
        },
    )


# Configurar el logger
logger = logging.getLogger(__name__)





@login_required
@transaction.atomic
def stock_ajuste(request):
    if request.method != "POST":
        return HttpResponseBadRequest("Método no permitido")

    try:
        pid = int(request.POST.get("producto_id"))
    except (TypeError, ValueError):
        return _json_or_redirect(request, False, "Producto inválido.")

    try:
        cantidad = float(request.POST.get("cantidad"))
    except (TypeError, ValueError):
        return _json_or_redirect(request, False, "Cantidad inválida.")

    ubicacion_id = request.POST.get("ubicacion")  # puede ser de bodega o de sucursal
    producto = get_object_or_404(Producto, pk=pid)

    if ubicacion_id:
        stock, _, _, _ = _get_or_create_stock(producto, int(ubicacion_id))
        stock.cantidad_disponible = F("cantidad_disponible") + Decimal(cantidad)
        stock.save(update_fields=["cantidad_disponible"])
    else:
        # sin ubicación → directo al producto
        producto.stock = max(0, int((producto.stock or 0) + cantidad))
        producto.save(update_fields=["stock"])

    _recalcular_stock_global(producto)

    return _json_or_redirect(
        request,
        True,
        f"Ajuste aplicado: {cantidad:+g} {producto.sku}",
        redirect_to=reverse("products"),
    )


@login_required
@transaction.atomic
def stock_entrada(request):
    if request.method != "POST":
        return HttpResponseBadRequest("Método no permitido")

    try:
        pid = int(request.POST.get("producto_id"))
    except (TypeError, ValueError):
        return _json_or_redirect(request, False, "Producto inválido.")

    try:
        cantidad = float(request.POST.get("cantidad"))
    except (TypeError, ValueError):
        return _json_or_redirect(request, False, "Cantidad inválida.")

    if cantidad <= 0:
        return _json_or_redirect(request, False, "La cantidad debe ser mayor a 0.")

    ubicacion_id = request.POST.get("ubicacion")
    producto = get_object_or_404(Producto, pk=pid)

    if ubicacion_id:
        stock, _, _, _ = _get_or_create_stock(producto, int(ubicacion_id))
        stock.cantidad_disponible = F("cantidad_disponible") + Decimal(cantidad)
        stock.save(update_fields=["cantidad_disponible"])
    else:
        producto.stock = max(0, int((producto.stock or 0) + cantidad))
        producto.save(update_fields=["stock"])

    _recalcular_stock_global(producto)

    return _json_or_redirect(
        request,
        True,
        f"Entrada registrada (+{cantidad:g}) para {producto.sku}",
        redirect_to=reverse("products"),
    )


@login_required
@transaction.atomic
def stock_transferir(request):
    """
    Transfiere stock entre ubicaciones (pueden ser de tablas distintas).
    """
    if request.method != "POST":
        return HttpResponseBadRequest("Método no permitido")

    try:
        pid = int(request.POST.get("producto_id"))
        origen = int(request.POST.get("origen"))
        destino = int(request.POST.get("destino"))
    except (TypeError, ValueError):
        return _json_or_redirect(request, False, "Parámetros inválidos.")

    if origen == destino:
        return _json_or_redirect(request, False, "Origen y destino no pueden ser iguales.")

    try:
        cantidad = float(request.POST.get("cantidad"))
    except (TypeError, ValueError):
        return _json_or_redirect(request, False, "Cantidad inválida.")

    if cantidad <= 0:
        return _json_or_redirect(request, False, "La cantidad debe ser mayor a 0.")

    producto = get_object_or_404(Producto, pk=pid)

    stock_origen, _, _, _ = _get_or_create_stock(producto, origen)
    stock_destino, _, _, _ = _get_or_create_stock(producto, destino)

    # necesitamos el valor real
    stock_origen.refresh_from_db()

    if stock_origen.cantidad_disponible < cantidad:
        return _json_or_redirect(request, False, "Stock insuficiente en origen.")

    stock_origen.cantidad_disponible = Decimal(stock_origen.cantidad_disponible) - Decimal(cantidad)
    stock_origen.save(update_fields=["cantidad_disponible"])

    stock_destino.cantidad_disponible = Decimal(stock_destino.cantidad_disponible) + Decimal(cantidad)
    stock_destino.save(update_fields=["cantidad_disponible"])

    _recalcular_stock_global(producto)

    return _json_or_redirect(
        request,
        True,
        f"Transferidos {cantidad:g} de {producto.sku}",
        redirect_to=reverse("products"),
    )


@login_required
@transaction.atomic
def stock_recuento(request):
    if request.method != "POST":
        return HttpResponseBadRequest("Método no permitido")

    try:
        pid = int(request.POST.get("producto_id"))
    except (TypeError, ValueError):
        return _json_or_redirect(request, False, "Producto inválido.")

    try:
        cantidad_real = float(request.POST.get("cantidad_real"))
    except (TypeError, ValueError):
        return _json_or_redirect(request, False, "Cantidad inválida.")

    ubicacion_id = request.POST.get("ubicacion")
    producto = get_object_or_404(Producto, pk=pid)

    if ubicacion_id:
        stock, _, _, _ = _get_or_create_stock(producto, int(ubicacion_id))
        stock.cantidad_disponible = Decimal(cantidad_real)
        stock.save(update_fields=["cantidad_disponible"])
    else:
        producto.stock = max(0, int(cantidad_real))
        producto.save(update_fields=["stock"])

    _recalcular_stock_global(producto)

    return _json_or_redirect(
        request,
        True,
        f"Recuento guardado ({cantidad_real:g}) para {producto.sku}",
        redirect_to=reverse("products"),
    )


def _recalcular_stock_global(producto: Producto) -> None:
    """
    Suma TODO el stock por ubicaciones y lo deja en producto.stock.
    Evita el error de 'mixed types' forzando DecimalField.
    """
    agg = (
        Stock.objects
        .filter(producto=producto)
        .aggregate(
            total=Coalesce(
                Sum("cantidad_disponible"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            )
        )
    )
    total = agg["total"] or 0
    producto.stock = int(total)
    producto.save(update_fields=["stock"])



@require_GET
def ajax_sucursales_y_productos(request):
    bodega_id = request.GET.get("bodega_id")
    if not bodega_id:
        return JsonResponse({"error": "Falta bodega_id"}, status=400)

    try:
        bodega = Bodega.objects.get(pk=bodega_id)
    except Bodega.DoesNotExist:
        return JsonResponse({"error": "Bodega no encontrada"}, status=404)

    # ===== Sucursales de la bodega =====
    sucursales = list(
        bodega.sucursales.all()
        .order_by("codigo")
        .values("id", "codigo", "nombre")
    )

    # ===== Productos: SOLO con stock validado y stock disponible en ESTA bodega =====
    productos_qs = (
        Producto.objects
        .filter(
            stocks__ubicacion_bodega__bodega=bodega,
            stocks__cantidad_disponible__gt=0,   # stock físico > 0
            stocks__cantidad_validada__gt=0,     # stock validado > 0
        )
        .annotate(
            stock_total=Coalesce(
                Sum("stocks__cantidad_disponible"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            ),
            stock_validado=Coalesce(
                Sum("stocks__cantidad_validada"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            ),
        )
        .values("id", "sku", "nombre", "stock_total", "stock_validado")
        .distinct()
    )

    return JsonResponse(
        {
            "sucursales": sucursales,
            "productos": list(productos_qs),
        }
    )
def _resolver_ubicacion(pk: int):
    """
    Recibe un PK y trata de adivinar si es una ubicacion de bodega o de sucursal.
    Devuelve ('bodega', obj) o ('sucursal', obj).
    Lanza Http404 si no existe.
    """
    try:
        ub_bod = UbicacionBodega.objects.get(pk=pk)
        return "bodega", ub_bod
    except UbicacionBodega.DoesNotExist:
        pass
    try:
        ub_suc = UbicacionSucursal.objects.get(pk=pk)
        return "sucursal", ub_suc
    except UbicacionSucursal.DoesNotExist:
        raise Http404("Ubicación no encontrada")


def _get_or_create_stock(producto: Producto, ubicacion_pk: int | None = None):
    """
    Devuelve (stock, creado, tipo, ubicacion_instance)
    - si viene ubicacion_pk, intenta en bodega, luego sucursal
    - si NO viene ubicacion_pk → trabajamos solo en producto.stock (compat)
    """
    if ubicacion_pk is None:
        # modo compat: no hay ubicación → operamos en el campo producto.stock
        # devolvemos un "fake" con la misma API que Stock
        class _Proxy:
            def __init__(self, prod):
                self._p = prod

            @property
            def cantidad_disponible(self):
                return Decimal(self._p.stock or 0)

            @cantidad_disponible.setter
            def cantidad_disponible(self, val):
                self._p.stock = int(Decimal(val or 0))
                self._p.save(update_fields=["stock"])

            def save(self, update_fields=None):
                self._p.save(update_fields=["stock"])

        return _Proxy(producto), False, "producto", producto

    tipo, ubi = _resolver_ubicacion(ubicacion_pk)
    if tipo == "bodega":
        stock, created = Stock.objects.get_or_create(
            producto=producto,
            ubicacion_bodega=ubi,
            defaults={"cantidad_disponible": 0},
        )
        return stock, created, tipo, ubi
    else:
        stock, created = Stock.objects.get_or_create(
            producto=producto,
            ubicacion_sucursal=ubi,
            defaults={"cantidad_disponible": 0},
        )
        return stock, created, tipo, ubi


def _recalcular_stock_global(producto: Producto) -> None:
    """
    Suma TODO el stock por ubicaciones (bodega + sucursal) y lo deja en producto.stock.
    """
    agg = (
        Stock.objects.filter(producto=producto)
        .aggregate(
            total=Coalesce(
                Sum("cantidad_disponible"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            )
        )
    )
    total = agg["total"] or 0
    producto.stock = int(total)
    producto.save(update_fields=["stock"])






@login_required
def bodega_a_sucursal(request):
    # Para desplegar siempre los combos
    bodegas = (
        Bodega.objects
        .prefetch_related("sucursales", "ubicaciones")
        .order_by("codigo")
    )

    if request.method == "POST":
        bodega_id    = request.POST.get("bodega")
        sucursal_id  = request.POST.get("sucursal")
        producto_id  = request.POST.get("producto")
        cantidad_raw = request.POST.get("cantidad")

        # ===== Validaciones básicas del formulario =====
        if not all([bodega_id, sucursal_id, producto_id, cantidad_raw]):
            return render(
                request,
                "core/Movimientos/bodega_a_sucursal.html",
                {"bodegas": bodegas, "error": "Faltan datos en el formulario."},
            )

        # cantidad como entero primero
        try:
            cantidad_int = int(cantidad_raw)
        except (TypeError, ValueError):
            return render(
                request,
                "core/Movimientos/bodega_a_sucursal.html",
                {"bodegas": bodegas, "error": "La cantidad debe ser un número entero."},
            )

        if cantidad_int <= 0:
            return render(
                request,
                "core/Movimientos/bodega_a_sucursal.html",
                {"bodegas": bodegas, "error": "La cantidad debe ser mayor que 0."},
            )

        # luego la usamos como Decimal para todo el cálculo
        cantidad = Decimal(cantidad_int)

        # ===== Instancias firmes =====
        try:
            bodega   = Bodega.objects.get(pk=bodega_id)
            sucursal = Sucursal.objects.get(pk=sucursal_id, bodega=bodega)
            producto = Producto.objects.get(pk=producto_id)
        except (Bodega.DoesNotExist, Sucursal.DoesNotExist, Producto.DoesNotExist):
            return render(
                request,
                "core/Movimientos/bodega_a_sucursal.html",
                {"bodegas": bodegas, "error": "Alguna entidad seleccionada no existe."},
            )

        # Ubicación "origen" de referencia (solo para el Kárdex)
        ubi_origen = bodega.ubicaciones.first()

        # Ubicación destino en sucursal: siempre la default "SIN UBICACIÓN"
        ubi_destino = ensure_ubicacion_sucursal(sucursal)

        if not ubi_origen or not ubi_destino:
            return render(
                request,
                "core/Movimientos/bodega_a_sucursal.html",
                {
                    "bodegas": bodegas,
                    "error": "Faltan ubicaciones en bodega o sucursal.",
                },
            )

        transferencia = None

        # ===== Movimiento de stock + Transferencia =====
        with transaction.atomic():
            # 1) Filas de stock en ESA bodega con cantidad_validada > 0
            filas_origen = list(
                Stock.objects
                .select_for_update()
                .filter(
                    producto=producto,
                    ubicacion_bodega__bodega=bodega,
                    cantidad_disponible__gt=0,
                    cantidad_validada__gt=0,           # 👈 SOLO lo validado
                )
                .order_by("-cantidad_validada")        # primero las que más tienen validado
            )

            # Cupo total **validado** en la bodega
            total_validado = (
                Stock.objects
                .filter(
                    producto=producto,
                    ubicacion_bodega__bodega=bodega,
                )
                .aggregate(
                    t=Coalesce(
                        Sum("cantidad_validada"),
                        Value(
                            0,
                            output_field=DecimalField(max_digits=20, decimal_places=6),
                        ),
                    )
                )["t"]
                or 0
            )

            if total_validado < cantidad:
                return render(
                    request,
                    "core/Movimientos/bodega_a_sucursal.html",
                    {
                        "bodegas": bodegas,
                        "error": (
                            "No hay stock validado suficiente para mover desde esta bodega. "
                            f"Validado para movimiento: {total_validado}."
                        ),
                    },
                )

            # 2) CABECERA Transferencia: BODEGA → SUCURSAL
            transferencia = Transferencia.objects.create(
                tipo_movimiento="BOD_SUC",
                bodega_origen=bodega,
                bodega_destino=None,
                sucursal_origen=None,
                sucursal_destino=sucursal,
                estado="CONFIRMADA",
                creado_por=request.user if request.user.is_authenticated else None,
            )

            # 3) Destino: sucursal (stock en ubicación destino "SIN UBICACIÓN")
            stock_destino, _ = (
                Stock.objects
                .select_for_update()
                .get_or_create(
                    producto=producto,
                    ubicacion_sucursal=ubi_destino,
                    defaults={"cantidad_disponible": Decimal("0")},
                )
            )

            # 4) Repartimos la salida SOLO sobre lo validado
            faltante = cantidad

            for fila in filas_origen:
                if faltante <= 0:
                    break

                # Lo máximo que puedo mover de esta fila es lo validado y disponible
                max_de_esta_fila = min(fila.cantidad_validada, fila.cantidad_disponible)
                mueve = min(max_de_esta_fila, faltante)
                if mueve <= 0:
                    continue

                fila.cantidad_disponible = F("cantidad_disponible") - mueve
                fila.cantidad_validada   = F("cantidad_validada")   - mueve
                fila.save(update_fields=["cantidad_disponible", "cantidad_validada"])

                stock_destino.cantidad_disponible = F("cantidad_disponible") + mueve
                faltante -= mueve

            stock_destino.save(update_fields=["cantidad_disponible"])

            # 5) LÍNEA de transferencia
            LineaTransferencia.objects.create(
                transferencia=transferencia,
                producto=producto,
                lote=None,
                serie=None,
                cantidad=cantidad,
                unidad=_unidad_default(),
            )

            # 6) KÁRDEX (mov. de stock)
            try:
                MovimientoStock.objects.create(
                    tipo_movimiento=_tm("TRANSFERENCIA"),
                    producto=producto,
                    ubicacion_bodega_desde=ubi_origen,
                    ubicacion_sucursal_desde=None,
                    ubicacion_bodega_hasta=None,
                    ubicacion_sucursal_hasta=ubi_destino,
                    lote=None,
                    serie=None,
                    cantidad=cantidad,
                    unidad=_unidad_default(),
                    tabla_referencia="transferencias",
                    referencia_id=transferencia.id,
                    creado_por=request.user if request.user.is_authenticated else None,
                    notas=f"Bodega {bodega.codigo} → Sucursal {sucursal.codigo}",
                )
            except TipoMovimiento.DoesNotExist:
                # Si aún no existe el TipoMovimiento, simplemente no grabamos en Kárdex
                pass

        # ===== Recalcular stock global del producto =====
        total_prod = (
            Stock.objects
            .filter(producto=producto)
            .aggregate(
                t=Coalesce(
                    Sum("cantidad_disponible"),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                )
            )["t"]
            or 0
        )
        producto.stock = int(total_prod)
        producto.save(update_fields=["stock"])

        # ===== Stock restante en ESA bodega + alerta de stock bajo =====
        rem_total_bodega = (
            Stock.objects
            .filter(
                producto=producto,
                ubicacion_bodega__bodega=bodega,
            )
            .aggregate(
                t=Coalesce(
                    Sum("cantidad_disponible"),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                )
            )["t"]
            or 0
        )

        notificar_stock_bajo(
            producto=producto,
            nombre_lugar=bodega.nombre,
            stock_actual=rem_total_bodega,
        )

        # ===== Recargar combos para la vista =====
        sucursales_de_bodega = bodega.sucursales.all().order_by("codigo")
        productos_de_bodega = (
            Producto.objects
            # Solo filtramos por bodega
            .filter(
                stocks__ubicacion_bodega__bodega=bodega,
            )
            .annotate(
                # STOCK FÍSICO REAL en esa bodega
                stock_total=Coalesce(
                    Sum(
                        "stocks__cantidad_disponible",
                        filter=Q(stocks__ubicacion_bodega__bodega=bodega),
                    ),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                ),
                # STOCK VALIDADO en esa bodega
                stock_validado=Coalesce(
                    Sum(
                        "stocks__cantidad_validada",
                        filter=Q(stocks__ubicacion_bodega__bodega=bodega),
                    ),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                ),
            )
            # Mostramos solo productos con algo validado y algo de stock
            .filter(
                stock_validado__gt=0,
                stock_total__gt=0,
            )
            .distinct()
        )

        context = {
            "bodegas": bodegas,
            "sucursales": sucursales_de_bodega,
            "productos": productos_de_bodega,
            "success": (
                f"Movimiento realizado: {cantidad_int} de "
                f"{producto.nombre} → {sucursal.nombre}."
            ),
            "bodega_sel": bodega.id,
        }

        if transferencia is not None:
            context["transferencia_id"] = transferencia.id

        return render(
            request,
            "core/Movimientos/bodega_a_sucursal.html",
            context,
        )

    # ===== GET =====
    return render(
        request,
        "core/Movimientos/bodega_a_sucursal.html",
        {"bodegas": bodegas},
    )

@login_required
def sucursal_a_bodega(request):
    # Todas las bodegas y sucursales base
    bodegas_all = (
        Bodega.objects
        .prefetch_related("sucursales", "ubicaciones")
        .order_by("codigo")
    )
    sucursales_all = Sucursal.objects.order_by("codigo")

    if request.method == "POST":
        sucursal_id  = request.POST.get("sucursal")
        bodega_id    = request.POST.get("bodega")
        producto_id  = request.POST.get("producto")
        cantidad_raw = request.POST.get("cantidad")

        # ========== POST PARCIAL 0: SOLO SUCURSAL ==========
        if sucursal_id and not any([bodega_id, producto_id, cantidad_raw]):
            bodegas_de_sucursal = (
                bodegas_all.filter(sucursales__id=sucursal_id).distinct()
            )
            return render(
                request,
                "core/Movimientos/sucursal_a_bodega.html",
                {
                    "bodegas": bodegas_de_sucursal,
                    "sucursales": sucursales_all,
                    "productos": None,
                    "sucursal_sel": sucursal_id,
                },
            )

        # ========== POST PARCIAL 1: SUCURSAL + BODEGA ==========
        if sucursal_id and bodega_id and not all([producto_id, cantidad_raw]):
            try:
                bodega   = Bodega.objects.get(pk=bodega_id)
                sucursal = Sucursal.objects.get(pk=sucursal_id, bodega=bodega)
            except (Bodega.DoesNotExist, Sucursal.DoesNotExist):
                bodegas_de_sucursal = (
                    bodegas_all.filter(sucursales__id=sucursal_id).distinct()
                    if sucursal_id
                    else bodegas_all
                )
                return render(
                    request,
                    "core/Movimientos/sucursal_a_bodega.html",
                    {
                        "bodegas": bodegas_de_sucursal,
                        "sucursales": sucursales_all,
                        "error": "La bodega o la sucursal seleccionada no existe.",
                        "sucursal_sel": sucursal_id,
                    },
                )

            bodegas_de_sucursal = (
                bodegas_all.filter(sucursales__id=sucursal.id).distinct()
            )

            # SOLO productos con stock disponible y VALIDADO en la sucursal
            productos_de_sucursal = (
                Producto.objects
                .filter(
                    stocks__ubicacion_sucursal__sucursal=sucursal,
                    stocks__cantidad_disponible__gt=0,
                    stocks__cantidad_validada__gt=0,
                )
                .annotate(
                    stock_total=Coalesce(
                        Sum(
                            "stocks__cantidad_disponible",
                            filter=Q(stocks__ubicacion_sucursal__sucursal=sucursal),
                        ),
                        Value(
                            0,
                            output_field=DecimalField(max_digits=20, decimal_places=6),
                        ),
                    ),
                    stock_validado=Coalesce(
                        Sum(
                            "stocks__cantidad_validada",
                            filter=Q(stocks__ubicacion_sucursal__sucursal=sucursal),
                        ),
                        Value(
                            0,
                            output_field=DecimalField(max_digits=20, decimal_places=6),
                        ),
                    ),
                )
                .distinct()
            )

            return render(
                request,
                "core/Movimientos/sucursal_a_bodega.html",
                {
                    "bodegas": bodegas_de_sucursal,
                    "sucursales": sucursales_all,
                    "productos": productos_de_sucursal,
                    "bodega_sel": bodega.id,
                    "sucursal_sel": sucursal.id,
                },
            )

        # ========== POST COMPLETO ==========
        if not all([sucursal_id, bodega_id, producto_id, cantidad_raw]):
            bodegas_ctx = (
                bodegas_all.filter(sucursales__id=sucursal_id).distinct()
                if sucursal_id
                else bodegas_all
            )
            return render(
                request,
                "core/Movimientos/sucursal_a_bodega.html",
                {
                    "bodegas": bodegas_ctx,
                    "sucursales": sucursales_all,
                    "error": "Faltan datos en el formulario.",
                    "sucursal_sel": sucursal_id,
                    "bodega_sel": bodega_id,
                },
            )

        try:
            cantidad_int = int(cantidad_raw)
        except (ValueError, TypeError):
            bodegas_ctx = (
                bodegas_all.filter(sucursales__id=sucursal_id).distinct()
                if sucursal_id
                else bodegas_all
            )
            return render(
                request,
                "core/Movimientos/sucursal_a_bodega.html",
                {
                    "bodegas": bodegas_ctx,
                    "sucursales": sucursales_all,
                    "error": "La cantidad debe ser un número entero.",
                    "sucursal_sel": sucursal_id,
                    "bodega_sel": bodega_id,
                },
            )

        if cantidad_int <= 0:
            bodegas_ctx = (
                bodegas_all.filter(sucursales__id=sucursal_id).distinct()
                if sucursal_id
                else bodegas_all
            )
            return render(
                request,
                "core/Movimientos/sucursal_a_bodega.html",
                {
                    "bodegas": bodegas_ctx,
                    "sucursales": sucursales_all,
                    "error": "La cantidad debe ser mayor que 0.",
                    "sucursal_sel": sucursal_id,
                    "bodega_sel": bodega_id,
                },
            )

        cantidad_dec = Decimal(cantidad_int)

        try:
            bodega   = Bodega.objects.get(pk=bodega_id)
            sucursal = Sucursal.objects.get(pk=sucursal_id, bodega=bodega)
            producto = Producto.objects.get(pk=producto_id)
        except (Bodega.DoesNotExist, Sucursal.DoesNotExist, Producto.DoesNotExist):
            bodegas_ctx = (
                bodegas_all.filter(sucursales__id=sucursal_id).distinct()
                if sucursal_id
                else bodegas_all
            )
            return render(
                request,
                "core/Movimientos/sucursal_a_bodega.html",
                {
                    "bodegas": bodegas_ctx,
                    "sucursales": sucursales_all,
                    "error": "Alguna entidad seleccionada no existe.",
                    "sucursal_sel": sucursal_id,
                    "bodega_sel": bodega_id,
                },
            )

        # Ubicaciones
        try:
            ubi_origen = ensure_ubicacion_sucursal(sucursal)
        except Exception:
            ubi_origen = sucursal.ubicaciones.first()

        ubi_destino = bodega.ubicaciones.first()

        if not ubi_origen or not ubi_destino:
            bodegas_ctx = (
                bodegas_all.filter(sucursales__id=sucursal.id).distinct()
            )
            return render(
                request,
                "core/Movimientos/sucursal_a_bodega.html",
                {
                    "bodegas": bodegas_ctx,
                    "sucursales": sucursales_all,
                    "error": "Faltan ubicaciones en la sucursal o en la bodega.",
                    "sucursal_sel": sucursal.id,
                    "bodega_sel": bodega.id,
                },
            )

        transferencia = None

        with transaction.atomic():
            # ORIGEN: SUCURSAL, SOLO filas con stock VALIDADO
            filas_origen = list(
                Stock.objects
                .select_for_update()
                .filter(
                    producto=producto,
                    ubicacion_sucursal__sucursal=sucursal,
                    cantidad_disponible__gt=0,
                    cantidad_validada__gt=0,
                )
                .order_by("-cantidad_validada")
            )

            # Cupo total VALIDADO en la sucursal
            total_validado = (
                Stock.objects
                .filter(
                    producto=producto,
                    ubicacion_sucursal__sucursal=sucursal,
                )
                .aggregate(
                    t=Coalesce(
                        Sum("cantidad_validada"),
                        Value(
                            0,
                            output_field=DecimalField(
                                max_digits=20,
                                decimal_places=6,
                            ),
                        ),
                    )
                )["t"]
                or Decimal("0")
            )

            if total_validado < cantidad_dec:
                bodegas_ctx = (
                    bodegas_all.filter(sucursales__id=sucursal.id).distinct()
                )
                return render(
                    request,
                    "core/Movimientos/sucursal_a_bodega.html",
                    {
                        "bodegas": bodegas_ctx,
                        "sucursales": sucursales_all,
                        "error": (
                            "No hay stock validado suficiente en la sucursal. "
                            f"Validado: {total_validado}."
                        ),
                        "sucursal_sel": sucursal.id,
                        "bodega_sel": bodega.id,
                    },
                )

            # DESTINO: BODEGA (stock en ubicación destino)
            stock_destino, _ = (
                Stock.objects
                .select_for_update()
                .get_or_create(
                    producto=producto,
                    ubicacion_bodega=ubi_destino,
                    defaults={"cantidad_disponible": Decimal("0")},
                )
            )

            # Repartimos la salida SOLO sobre lo validado en sucursal
            faltante = cantidad_dec

            for fila in filas_origen:
                if faltante <= 0:
                    break

                max_de_esta_fila = min(fila.cantidad_validada, fila.cantidad_disponible)
                mueve = min(max_de_esta_fila, faltante)
                if mueve <= 0:
                    continue

                fila.cantidad_disponible = fila.cantidad_disponible - mueve
                fila.cantidad_validada   = fila.cantidad_validada   - mueve
                fila.save(update_fields=["cantidad_disponible", "cantidad_validada"])

                stock_destino.cantidad_disponible = (
                    stock_destino.cantidad_disponible + mueve
                )
                faltante -= mueve

            stock_destino.save(update_fields=["cantidad_disponible"])

            # CABECERA: SUCURSAL → BODEGA
            transferencia = Transferencia.objects.create(
                tipo_movimiento="SUC_BOD",
                bodega_origen=None,
                bodega_destino=bodega,
                sucursal_origen=sucursal,
                sucursal_destino=None,
                estado="CONFIRMADA",
                creado_por=request.user if request.user.is_authenticated else None,
            )

            # LÍNEA
            LineaTransferencia.objects.create(
                transferencia=transferencia,
                producto=producto,
                lote=None,
                serie=None,
                cantidad=cantidad_dec,
                unidad=_unidad_default(),
            )

            # KÁRDEX
            try:
                MovimientoStock.objects.create(
                    tipo_movimiento=_tm("TRANSFERENCIA"),
                    producto=producto,
                    ubicacion_bodega_desde=None,
                    ubicacion_sucursal_desde=ubi_origen,
                    ubicacion_bodega_hasta=ubi_destino,
                    ubicacion_sucursal_hasta=None,
                    lote=None,
                    serie=None,
                    cantidad=cantidad_dec,
                    unidad=_unidad_default(),
                    tabla_referencia="transferencias",
                    referencia_id=transferencia.id,
                    creado_por=request.user if request.user.is_authenticated else None,
                    notas=f"Suc {sucursal.codigo} → Bod {bodega.codigo}",
                )
            except TipoMovimiento.DoesNotExist:
                pass

        # Recalcular stock global
        total_prod = (
            Stock.objects
            .filter(producto=producto)
            .aggregate(
                t=Coalesce(
                    Sum("cantidad_disponible"),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                )
            )["t"]
            or Decimal("0")
        )
        producto.stock = int(total_prod)
        producto.save(update_fields=["stock"])

        # Stock restante en sucursal (todas las ubicaciones de esa sucursal)
        rem_total_sucursal = (
            Stock.objects
            .filter(
                producto=producto,
                ubicacion_sucursal__sucursal=sucursal,
            )
            .aggregate(
                t=Coalesce(
                    Sum("cantidad_disponible"),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                )
            )["t"]
            or Decimal("0")
        )

        notificar_stock_bajo(
            producto=producto,
            nombre_lugar=sucursal.nombre,
            stock_actual=rem_total_sucursal,
        )

        # Recargar combos coherentes (SOLO validados)
        bodegas_de_sucursal = bodegas_all.filter(sucursales__id=sucursal.id).distinct()
        productos_de_sucursal = (
            Producto.objects
            .filter(
                stocks__ubicacion_sucursal__sucursal=sucursal,
                stocks__cantidad_disponible__gt=0,
                stocks__cantidad_validada__gt=0,
            )
            .annotate(
                stock_total=Coalesce(
                    Sum(
                        "stocks__cantidad_disponible",
                        filter=Q(stocks__ubicacion_sucursal__sucursal=sucursal),
                    ),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                ),
                stock_validado=Coalesce(
                    Sum(
                        "stocks__cantidad_validada",
                        filter=Q(stocks__ubicacion_sucursal__sucursal=sucursal),
                    ),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                ),
            )
            .distinct()
        )

        context = {
            "bodegas": bodegas_de_sucursal,
            "sucursales": sucursales_all,
            "productos": productos_de_sucursal,
            "success": (
                f"Movimiento realizado: {cantidad_int} de "
                f"{producto.nombre} → {bodega.nombre}."
            ),
            "bodega_sel": bodega.id,
            "sucursal_sel": sucursal.id,
        }

        if transferencia is not None:
            context["transferencia_id"] = transferencia.id

        return render(request, "core/Movimientos/sucursal_a_bodega.html", context)

    # GET simple
    return render(
        request,
        "core/Movimientos/sucursal_a_bodega.html",
        {
            "bodegas": bodegas_all,
            "sucursales": sucursales_all,
            "productos": None,
        },
    )



@login_required
def sucursal_a_sucursal(request):
    # 1) Selects base
    sucursales = (
        Sucursal.objects
        .prefetch_related("ubicaciones")
        .order_by("nombre")
    )

    # 2) Origen puede venir por GET (autosubmit) o POST
    suc_origen_id = request.GET.get("sucursal_origen") or request.POST.get("sucursal_origen")

    sucursales_destino = None
    productos_de_origen = None
    suc_origen = None

    # ========= Precarga cuando hay origen =========
    if suc_origen_id:
        try:
            suc_origen = Sucursal.objects.get(pk=suc_origen_id)
        except Sucursal.DoesNotExist:
            suc_origen = None
        else:
            # asegura que exista al menos 1 ubicación (SIN-UBI si no hay nada)
            try:
                ensure_ubicacion_sucursal(suc_origen)
            except Exception:
                pass

            sucursales_destino = (
                Sucursal.objects
                .exclude(pk=suc_origen.pk)
                .order_by("nombre")
            )

            # SOLO productos con stock disponible y VALIDADO en la sucursal origen
            productos_de_origen = (
                Producto.objects
                .filter(
                    stocks__ubicacion_sucursal__sucursal=suc_origen,
                    stocks__cantidad_disponible__gt=0,
                    stocks__cantidad_validada__gt=0,
                )
                .annotate(
                    stock_total=Coalesce(
                        Sum(
                            "stocks__cantidad_disponible",
                            filter=Q(stocks__ubicacion_sucursal__sucursal=suc_origen),
                        ),
                        Value(
                            0,
                            output_field=DecimalField(max_digits=20, decimal_places=6),
                        ),
                    ),
                    stock_validado=Coalesce(
                        Sum(
                            "stocks__cantidad_validada",
                            filter=Q(stocks__ubicacion_sucursal__sucursal=suc_origen),
                        ),
                        Value(
                            0,
                            output_field=DecimalField(max_digits=20, decimal_places=6),
                        ),
                    ),
                )
                .distinct()
            )

    # ========= POST PARCIAL (solo cambió el origen por autosubmit) =========
    if (
        request.method == "POST"
        and "sucursal_origen" in request.POST
        and not all([
            request.POST.get("sucursal_destino"),
            request.POST.get("producto"),
            request.POST.get("cantidad"),
        ])
    ):
        return render(
            request,
            "core/Movimientos/sucursal_a_sucursal.html",
            {
                "sucursales": sucursales,
                "sucursal_sel": suc_origen.id if suc_origen else None,
                "sucursales_destino": sucursales_destino,
                "productos": productos_de_origen,
            },
        )

    # ========= POST COMPLETO → ejecutar movimiento =========
    if request.method == "POST":
        suc_destino_id = request.POST.get("sucursal_destino")
        producto_id    = request.POST.get("producto")
        cantidad_raw   = request.POST.get("cantidad")

        if not all([suc_origen_id, suc_destino_id, producto_id, cantidad_raw]):
            return render(
                request,
                "core/Movimientos/sucursal_a_sucursal.html",
                {
                    "sucursales": sucursales,
                    "sucursal_sel": suc_origen_id,
                    "sucursales_destino": sucursales_destino,
                    "productos": productos_de_origen,
                    "error": "Faltan datos en el formulario.",
                },
            )

        if suc_origen_id == suc_destino_id:
            return render(
                request,
                "core/Movimientos/sucursal_a_sucursal.html",
                {
                    "sucursales": sucursales,
                    "sucursal_sel": suc_origen_id,
                    "sucursales_destino": sucursales_destino,
                    "productos": productos_de_origen,
                    "error": "La sucursal de origen y destino no pueden ser la misma.",
                },
            )

        # Parseo de cantidad
        try:
            cantidad_int = int(cantidad_raw)
        except (ValueError, TypeError):
            cantidad_int = 0

        if cantidad_int <= 0:
            return render(
                request,
                "core/Movimientos/sucursal_a_sucursal.html",
                {
                    "sucursales": sucursales,
                    "sucursal_sel": suc_origen_id,
                    "sucursales_destino": sucursales_destino,
                    "productos": productos_de_origen,
                    "error": "La cantidad debe ser mayor que 0.",
                },
            )

        cantidad = Decimal(cantidad_int)

        suc_origen  = get_object_or_404(Sucursal, pk=suc_origen_id)
        suc_destino = get_object_or_404(Sucursal, pk=suc_destino_id)
        producto    = get_object_or_404(Producto, pk=producto_id)

        # Ubicaciones default (crea SIN-UBI si no existe)
        try:
            ubi_origen = ensure_ubicacion_sucursal(suc_origen)
        except Exception:
            ubi_origen = suc_origen.ubicaciones.first()

        try:
            ubi_destino = ensure_ubicacion_sucursal(suc_destino)
        except Exception:
            ubi_destino = suc_destino.ubicaciones.first()

        if not ubi_origen or not ubi_destino:
            return render(
                request,
                "core/Movimientos/sucursal_a_sucursal.html",
                {
                    "sucursales": sucursales,
                    "sucursal_sel": suc_origen_id,
                    "sucursales_destino": sucursales_destino,
                    "productos": productos_de_origen,
                    "error": "Faltan ubicaciones en sucursal origen o destino.",
                },
            )

        transferencia = None  # para pasar el id al contexto

        with transaction.atomic():
            # ===== ORIGEN: SOLO filas con stock VALIDADO en sucursal origen =====
            filas_origen = list(
                Stock.objects
                .select_for_update()
                .filter(
                    producto=producto,
                    ubicacion_sucursal__sucursal=suc_origen,
                    cantidad_disponible__gt=0,
                    cantidad_validada__gt=0,
                )
                .order_by("-cantidad_validada")
            )

            # Cupo total VALIDADO en sucursal origen
            total_validado = (
                Stock.objects
                .filter(
                    producto=producto,
                    ubicacion_sucursal__sucursal=suc_origen,
                )
                .aggregate(
                    t=Coalesce(
                        Sum("cantidad_validada"),
                        Value(
                            0,
                            output_field=DecimalField(max_digits=20, decimal_places=6),
                        ),
                    )
                )["t"]
                or Decimal("0")
            )

            if total_validado < cantidad:
                return render(
                    request,
                    "core/Movimientos/sucursal_a_sucursal.html",
                    {
                        "sucursales": sucursales,
                        "sucursal_sel": suc_origen_id,
                        "sucursales_destino": sucursales_destino,
                        "productos": productos_de_origen,
                        "error": (
                            "No hay stock validado suficiente en "
                            f"{suc_origen.nombre}. Validado: {total_validado}."
                        ),
                    },
                )

            # ===== DESTINO =====
            stock_destino, _ = (
                Stock.objects
                .select_for_update()
                .get_or_create(
                    producto=producto,
                    ubicacion_sucursal=ubi_destino,
                    defaults={"cantidad_disponible": Decimal("0")},
                )
            )

            # ====== CABECERA Transferencia ======
            transferencia = Transferencia.objects.create(
                tipo_movimiento="SUC_SUC",
                bodega_origen=None,
                bodega_destino=None,
                sucursal_origen=suc_origen,    # origen
                sucursal_destino=suc_destino,  # destino
                estado="CONFIRMADA",
                creado_por=request.user if request.user.is_authenticated else None,
            )

            # LÍNEA
            LineaTransferencia.objects.create(
                transferencia=transferencia,
                producto=producto,
                lote=None,
                serie=None,
                cantidad=cantidad,
                unidad=_unidad_default(),
            )

            # ===== Aplicar movimiento SOLO sobre lo VALIDADO =====
            faltante = cantidad

            for fila in filas_origen:
                if faltante <= 0:
                    break

                max_de_esta_fila = min(fila.cantidad_validada, fila.cantidad_disponible)
                mueve = min(max_de_esta_fila, faltante)
                if mueve <= 0:
                    continue

                fila.cantidad_disponible = fila.cantidad_disponible - mueve
                fila.cantidad_validada   = fila.cantidad_validada   - mueve
                fila.save(update_fields=["cantidad_disponible", "cantidad_validada"])

                stock_destino.cantidad_disponible = (
                    stock_destino.cantidad_disponible + mueve
                )
                faltante -= mueve

            stock_destino.save(update_fields=["cantidad_disponible"])

            # === KÁRDEX / MovimientoStock ===
            try:
                MovimientoStock.objects.create(
                    tipo_movimiento=_tm("TRANSFERENCIA"),
                    producto=producto,
                    ubicacion_bodega_desde=None,
                    ubicacion_sucursal_desde=ubi_origen,
                    ubicacion_bodega_hasta=None,
                    ubicacion_sucursal_hasta=ubi_destino,
                    lote=None,
                    serie=None,
                    cantidad=cantidad,
                    unidad=_unidad_default(),
                    tabla_referencia="transferencias",
                    referencia_id=transferencia.id,
                    creado_por=request.user if request.user.is_authenticated else None,
                    notas=f"Suc {suc_origen.codigo} → Suc {suc_destino.codigo}",
                )
            except TipoMovimiento.DoesNotExist:
                # Si no existe el tipo de movimiento, no rompemos el flujo
                pass

        # ========= Recalcular stock global del producto =========
        total_prod = (
            Stock.objects
            .filter(producto=producto)
            .aggregate(
                total=Coalesce(
                    Sum("cantidad_disponible"),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                )
            )["total"]
        ) or 0
        producto.stock = int(total_prod)
        producto.save(update_fields=["stock"])

        # ========= Alerta por correo si queda bajo cierto umbral en ORIGEN =========
        rem_total_origen = (
            Stock.objects
            .filter(
                producto=producto,
                ubicacion_sucursal__sucursal=suc_origen,
            )
            .aggregate(
                t=Coalesce(
                    Sum("cantidad_disponible"),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                )
            )["t"]
            or Decimal("0")
        )

        if rem_total_origen < 10:
            notificar_stock_bajo(
                producto=producto,
                nombre_lugar=suc_origen.nombre,
                stock_actual=rem_total_origen,
            )

        # ========= Recargar combos para seguir moviendo desde la misma sucursal =========
        productos_disponibles = (
            Producto.objects
            .filter(
                stocks__ubicacion_sucursal__sucursal=suc_origen,
                stocks__cantidad_disponible__gt=0,
                stocks__cantidad_validada__gt=0,
            )
            .annotate(
                stock_total=Coalesce(
                    Sum(
                        "stocks__cantidad_disponible",
                        filter=Q(stocks__ubicacion_sucursal__sucursal=suc_origen),
                    ),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                ),
                stock_validado=Coalesce(
                    Sum(
                        "stocks__cantidad_validada",
                        filter=Q(stocks__ubicacion_sucursal__sucursal=suc_origen),
                    ),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                ),
            )
            .distinct()
        )

        context = {
            "sucursales": sucursales,
            "sucursal_sel": suc_origen.id,
            "sucursales_destino": (
                Sucursal.objects
                .exclude(pk=suc_origen.id)
                .order_by("nombre")
            ),
            "productos": productos_disponibles,
            "success": (
                f"Movimiento realizado: {cantidad_int} de "
                f"{producto.nombre} → {suc_destino.nombre}."
            ),
        }

        # Para que el template pueda mostrar el botón "Ver guía de despacho"
        if transferencia is not None:
            context["transferencia_id"] = transferencia.id

        return render(
            request,
            "core/Movimientos/sucursal_a_sucursal.html",
            context,
        )

    # ========= GET normal =========
    return render(
        request,
        "core/Movimientos/sucursal_a_sucursal.html",
        {
            "sucursales": sucursales,
            "sucursal_sel": suc_origen.id if suc_origen else None,
            "sucursales_destino": sucursales_destino,
            "productos": productos_de_origen,
        },
    )

@login_required
def bodega_a_bodega(request):
    bodegas = (
        Bodega.objects
        .prefetch_related("ubicaciones")
        .order_by("codigo")
    )

    if request.method == "POST":
        bodega_origen_id  = request.POST.get("bodega_origen")
        bodega_destino_id = request.POST.get("bodega_destino")
        producto_id       = request.POST.get("producto")
        cantidad_raw      = request.POST.get("cantidad")

        # ====== POST PARCIAL: solo bodega origen (cargar productos) ======
        if bodega_origen_id and not all([bodega_destino_id, producto_id, cantidad_raw]):
            try:
                bodega_origen = Bodega.objects.get(pk=bodega_origen_id)
            except Bodega.DoesNotExist:
                bodega_origen = None
                productos_de_origen = None
                bodegas_destino = bodegas
            else:
                # SOLO productos con stock disponible y VALIDADO en esta bodega
                productos_de_origen = (
                    Producto.objects
                    .filter(
                        stocks__ubicacion_bodega__bodega=bodega_origen,
                        stocks__cantidad_disponible__gt=0,
                        stocks__cantidad_validada__gt=0,
                    )
                    .annotate(
                        stock_total=Coalesce(
                            Sum(
                                "stocks__cantidad_disponible",
                                filter=Q(stocks__ubicacion_bodega__bodega=bodega_origen),
                            ),
                            Value(
                                0,
                                output_field=DecimalField(
                                    max_digits=20,
                                    decimal_places=6,
                                ),
                            ),
                        ),
                        stock_validado=Coalesce(
                            Sum(
                                "stocks__cantidad_validada",
                                filter=Q(stocks__ubicacion_bodega__bodega=bodega_origen),
                            ),
                            Value(
                                0,
                                output_field=DecimalField(
                                    max_digits=20,
                                    decimal_places=6,
                                ),
                            ),
                        ),
                    )
                    .distinct()
                )
                bodegas_destino = bodegas.exclude(pk=bodega_origen.id)

            return render(
                request,
                "core/Movimientos/bodega_a_bodega.html",
                {
                    "bodegas": bodegas,
                    "bodega_sel": bodega_origen_id,
                    "productos": productos_de_origen,
                    "bodegas_destino": bodegas_destino,
                    "bodega_destino_sel": bodega_destino_id,
                },
            )

        # ===== Validaciones básicas (POST COMPLETO) =====
        if not all([bodega_origen_id, bodega_destino_id, producto_id, cantidad_raw]):
            return render(
                request,
                "core/Movimientos/bodega_a_bodega.html",
                {"bodegas": bodegas, "error": "Faltan datos en el formulario."},
            )

        if bodega_origen_id == bodega_destino_id:
            return render(
                request,
                "core/Movimientos/bodega_a_bodega.html",
                {
                    "bodegas": bodegas,
                    "error": "La bodega de origen y destino no pueden ser la misma.",
                },
            )

        # Parseo de cantidad
        try:
            cantidad = int(cantidad_raw)
        except (TypeError, ValueError):
            return render(
                request,
                "core/Movimientos/bodega_a_bodega.html",
                {"bodegas": bodegas, "error": "La cantidad debe ser un número entero."},
            )

        if cantidad <= 0:
            return render(
                request,
                "core/Movimientos/bodega_a_bodega.html",
                {"bodegas": bodegas, "error": "La cantidad debe ser mayor que 0."},
            )

        cantidad_dec = Decimal(cantidad)

        # Instancias firmes (solo para lógica)
        try:
            bodega_origen  = Bodega.objects.get(pk=bodega_origen_id)
            bodega_destino = Bodega.objects.get(pk=bodega_destino_id)
            producto       = Producto.objects.get(pk=producto_id)
        except (Bodega.DoesNotExist, Producto.DoesNotExist):
            return render(
                request,
                "core/Movimientos/bodega_a_bodega.html",
                {"bodegas": bodegas, "error": "Alguna entidad seleccionada no existe."},
            )

        # Ubicaciones
        ubi_origen  = bodega_origen.ubicaciones.first()
        ubi_destino = bodega_destino.ubicaciones.first()

        if not ubi_origen or not ubi_destino:
            return render(
                request,
                "core/Movimientos/bodega_a_bodega.html",
                {
                    "bodegas": bodegas,
                    "error": "Faltan ubicaciones en alguna bodega.",
                },
            )

        transferencia = None

        # ===== Movimiento de stock + Transferencia =====
        with transaction.atomic():
            # 1) Filas de stock ORIGEN solo con VALIDADO
            filas_origen = list(
                Stock.objects
                .select_for_update()
                .filter(
                    producto=producto,
                    ubicacion_bodega__bodega=bodega_origen,
                    cantidad_disponible__gt=0,
                    cantidad_validada__gt=0,   # 👈 SOLO validado
                )
                .order_by("-cantidad_validada")        # primero las que más validado tienen
            )

            # Cupo total VALIDADO en la bodega origen
            total_validado = (
                Stock.objects
                .filter(
                    producto=producto,
                    ubicacion_bodega__bodega=bodega_origen,
                )
                .aggregate(
                    t=Coalesce(
                        Sum("cantidad_validada"),
                        Value(
                            0,
                            output_field=DecimalField(
                                max_digits=20,
                                decimal_places=6,
                            ),
                        ),
                    )
                )["t"]
                or 0
            )

            if total_validado < cantidad_dec:
                return render(
                    request,
                    "core/Movimientos/bodega_a_bodega.html",
                    {
                        "bodegas": bodegas,
                        "error": (
                            "No hay stock validado suficiente en la bodega origen. "
                            f"Validado: {total_validado}."
                        ),
                    },
                )

            # 2) CABECERA de la transferencia
            transferencia = Transferencia.objects.create(
                tipo_movimiento="BOD_BOD",
                bodega_origen=bodega_origen,
                bodega_destino=bodega_destino,
                sucursal_origen=None,
                sucursal_destino=None,
                estado="CONFIRMADA",
                creado_por=request.user if request.user.is_authenticated else None,
            )

            # 3) Stock destino (en bodega destino)
            stock_destino, _ = (
                Stock.objects
                .select_for_update()
                .get_or_create(
                    producto=producto,
                    ubicacion_bodega=ubi_destino,
                    defaults={"cantidad_disponible": Decimal("0")},
                )
            )

            # 4) Repartir la salida SOLO sobre stock validado
            faltante = cantidad_dec

            for fila in filas_origen:
                if faltante <= 0:
                    break

                # Lo máximo que puedo mover de esta fila es lo validado y disponible
                max_de_esta_fila = min(fila.cantidad_validada, fila.cantidad_disponible)
                mueve = min(max_de_esta_fila, faltante)
                if mueve <= 0:
                    continue

                fila.cantidad_disponible = fila.cantidad_disponible - mueve
                fila.cantidad_validada   = fila.cantidad_validada - mueve
                fila.save(update_fields=["cantidad_disponible", "cantidad_validada"])

                stock_destino.cantidad_disponible = (
                    stock_destino.cantidad_disponible + mueve
                )
                faltante -= mueve

            stock_destino.save(update_fields=["cantidad_disponible"])

            # 5) Línea Transferencia
            LineaTransferencia.objects.create(
                transferencia=transferencia,
                producto=producto,
                lote=None,
                serie=None,
                cantidad=cantidad_dec,
                unidad=_unidad_default(),
            )

            # 6) Kardex
            try:
                MovimientoStock.objects.create(
                    tipo_movimiento=_tm("TRANSFERENCIA"),
                    producto=producto,
                    ubicacion_bodega_desde=ubi_origen,
                    ubicacion_sucursal_desde=None,
                    ubicacion_bodega_hasta=ubi_destino,
                    ubicacion_sucursal_hasta=None,
                    lote=None,
                    serie=None,
                    cantidad=cantidad_dec,
                    unidad=_unidad_default(),
                    tabla_referencia="transferencias",
                    referencia_id=transferencia.id,
                    creado_por=request.user if request.user.is_authenticated else None,
                    notas=f"Bodega {bodega_origen.codigo} → Bodega {bodega_destino.codigo}",
                )
            except TipoMovimiento.DoesNotExist:
                pass

        # ===== Recalcular stock global producto =====
        total_prod = (
            Stock.objects
            .filter(producto=producto)
            .aggregate(
                t=Coalesce(
                    Sum("cantidad_disponible"),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                )
            )["t"]
            or 0
        )
        producto.stock = int(total_prod)
        producto.save(update_fields=["stock"])

        # ===== Stock restante en bodega origen + alerta =====
        rem_total_origen = (
            Stock.objects
            .filter(
                producto=producto,
                ubicacion_bodega__bodega=bodega_origen,
            )
            .aggregate(
                t=Coalesce(
                    Sum("cantidad_disponible"),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                )
            )["t"]
            or 0
        )

        notificar_stock_bajo(
            producto=producto,
            nombre_lugar=bodega_origen.nombre,
            stock_actual=rem_total_origen,
        )

        # ===== Recargar combos (solo productos validados) =====
        productos_de_origen = (
            Producto.objects
            .filter(
                stocks__ubicacion_bodega__bodega=bodega_origen,
                stocks__cantidad_disponible__gt=0,
                stocks__cantidad_validada__gt=0,
            )
            .annotate(
                stock_total=Coalesce(
                    Sum(
                        "stocks__cantidad_disponible",
                        filter=Q(stocks__ubicacion_bodega__bodega=bodega_origen),
                    ),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                ),
                stock_validado=Coalesce(
                    Sum(
                        "stocks__cantidad_validada",
                        filter=Q(stocks__ubicacion_bodega__bodega=bodega_origen),
                    ),
                    Value(
                        0,
                        output_field=DecimalField(max_digits=20, decimal_places=6),
                    ),
                ),
            )
            .distinct()
        )

        bodegas_destino = bodegas.exclude(pk=bodega_origen.id)

        context = {
            "bodegas": bodegas,
            "bodega_sel": bodega_origen.id,
            "bodega_destino_sel": bodega_destino.id,
            "productos": productos_de_origen,
            "bodegas_destino": bodegas_destino,
            "success": (
                f"Movimiento realizado: {cantidad} de "
                f"{producto.nombre} → {bodega_destino.nombre}."
            ),
        }

        if transferencia is not None:
            context["transferencia_id"] = transferencia.id

        return render(
            request,
            "core/Movimientos/bodega_a_bodega.html",
            context,
        )

    # GET
    return render(
        request,
        "core/Movimientos/bodega_a_bodega.html",
        {"bodegas": bodegas},
    )




def movimientos_index(request):
    return render(request, "core/Movimientos/movimientos_index.html")



@require_GET
def geocode(request):
    q = (request.GET.get("q") or "").strip()
    if not q:
        return JsonResponse({"error": "missing query"}, status=400)

    url = "https://nominatim.openstreetmap.org/search"
    params = {
        "format": "json",
        "limit": 1,
        "q": q,
        "countrycodes": "cl",
    }
    headers = {
        # Nominatim pide User-Agent
        "User-Agent": "LogisticFour/1.0 (https://example.com)"
    }

    try:
        r = requests.get(url, params=params, headers=headers, timeout=5)
        r.raise_for_status()
    except requests.RequestException as e:
        return JsonResponse({"error": "upstream_error", "detail": str(e)}, status=502)

    data = r.json()
    if not data:
        return JsonResponse({"error": "not_found"}, status=404)

    item = data[0]
    return JsonResponse({
        "lat": item["lat"],
        "lon": item["lon"],
        "display_name": item.get("display_name", q),
    })
    
from django.utils import timezone
@require_POST
@login_required
@transaction.atomic
def paypal_stock_in(request):
    """
    Llega desde el JS de PayPal cuando el pago fue APROBADO.
    Hace TODO automático y si algo falla devuelve el error en JSON.
    """
    try:
        # -------------------------------------------------
        # 1) leer JSON
        # -------------------------------------------------
        try:
            data = json.loads(request.body.decode("utf-8"))
        except Exception as e:
            return JsonResponse({"ok": False, "error": f"JSON inválido: {e}"}, status=400)

        bodega_id = data.get("bodega_id")
        producto_id = data.get("producto_id")
        cantidad_raw = data.get("cantidad")
        paypal_id = data.get("paypal_id") or "SIN-ID"
        monto_usd_raw = data.get("monto_usd") or "0"

        if not bodega_id or not producto_id or not cantidad_raw:
            return JsonResponse({"ok": False, "error": "Faltan datos"}, status=400)

        # -------------------------------------------------
        # 2) cantidad válida
        # -------------------------------------------------
        try:
            cantidad = Decimal(str(cantidad_raw))
            if cantidad <= 0:
                raise ValueError
        except Exception:
            return JsonResponse({"ok": False, "error": "Cantidad inválida"}, status=400)

        # -------------------------------------------------
        # 3) buscar bodega y producto
        # -------------------------------------------------
        try:
            bodega = Bodega.objects.get(pk=bodega_id)
            producto = Producto.objects.get(pk=producto_id)
        except Bodega.DoesNotExist:
            return JsonResponse({"ok": False, "error": "Bodega no encontrada"}, status=404)
        except Producto.DoesNotExist:
            return JsonResponse({"ok": False, "error": "Producto no encontrado"}, status=404)

        # -------------------------------------------------
        # 4) asegurar UNA ubicación en esa bodega
        # -------------------------------------------------
        ubi = bodega.ubicaciones.filter(activo=True).order_by("id").first()
        if not ubi:
            ubi = UbicacionBodega.objects.create(
                bodega=bodega,
                codigo="AUTO-PP",
                nombre="Ubicación generada por PayPal",
                activo=True,
            )

        # -------------------------------------------------
        # 5) sumar stock en ESA ubicación
        # -------------------------------------------------
        stock_obj, created = Stock.objects.get_or_create(
            producto=producto,
            ubicacion_bodega=ubi,
            defaults={"cantidad_disponible": Decimal("0")}
        )
        Stock.objects.filter(pk=stock_obj.pk).update(
            cantidad_disponible=F("cantidad_disponible") + cantidad
        )
        stock_obj.refresh_from_db()

        # -------------------------------------------------
        # 6) usuario proveedor PayPal (rol = PROVEEDOR)
        # -------------------------------------------------
        proveedor_user, _ = User.objects.get_or_create(
            username="paypal_proveedor",
            defaults={
                "first_name": "Proveedor",
                "last_name": "PayPal",
                "email": "paypal@example.com",
            },
        )
        UsuarioPerfil.objects.get_or_create(
            usuario=proveedor_user,
            defaults={"rol": UsuarioPerfil.Rol.PROVEEDOR},
        )

        # -------------------------------------------------
        # 7) unidad de medida segura
        # -------------------------------------------------
        um = getattr(producto, "unidad_base", None)
        if not um:
            # intentamos agarrar una existente
            um = UnidadMedida.objects.first()
        if not um:
            # si no hay ninguna en la bd, creamos una por defecto
            um, _ = UnidadMedida.objects.get_or_create(
                codigo="UN-PP",
                defaults={"descripcion": "Unidad por PayPal"},
            )

        # -------------------------------------------------
        # 8) crear ORDEN DE COMPRA
        # -------------------------------------------------
        numero_orden = f"OC-PAYPAL-{timezone.now().strftime('%Y%m%d%H%M%S')}"
        oc = OrdenCompra.objects.create(
            proveedor=proveedor_user,
            tasa_impuesto=None,
            bodega=bodega,
            numero_orden=numero_orden,
            estado="COMPLETED",
            fecha_esperada=timezone.now().date(),
            creado_por=request.user,
        )

        # monto
        try:
            monto_usd = Decimal(str(monto_usd_raw))
        except Exception:
            monto_usd = Decimal("0")

        LineaOrdenCompra.objects.create(
            orden_compra=oc,
            producto=producto,
            descripcion=f"Compra PayPal #{paypal_id}",
            cantidad_pedida=cantidad,
            unidad=um,
            precio=monto_usd,
        )

        # -------------------------------------------------
        # 9) crear RECEPCIÓN
        # -------------------------------------------------
        rec = RecepcionMercaderia.objects.create(
            orden_compra=oc,
            bodega=bodega,
            numero_recepcion=f"RC-PAYPAL-{timezone.now().strftime('%Y%m%d%H%M%S')}",
            estado="CLOSED",
            recibido_por=request.user,
        )
        LineaRecepcionMercaderia.objects.create(
            recepcion=rec,
            producto=producto,
            cantidad_recibida=cantidad,
            unidad=um,
        )

        # -------------------------------------------------
        # 10) crear FACTURA (cuidando el unique)
        # -------------------------------------------------
        base_numero_factura = f"FP-PAYPAL-{paypal_id}"
        try:
            FacturaProveedor.objects.create(
                proveedor=proveedor_user,
                numero_factura=base_numero_factura,
                monto_total=monto_usd,
                fecha_factura=timezone.now().date(),
                estado="PAID",
            )
        except IntegrityError:
            # si ya existía una factura con ese número, le metemos sufijo de tiempo
            FacturaProveedor.objects.create(
                proveedor=proveedor_user,
                numero_factura=f"{base_numero_factura}-{timezone.now().strftime('%H%M%S')}",
                monto_total=monto_usd,
                fecha_factura=timezone.now().date(),
                estado="PAID",
            )

        # -------------------------------------------------
        # LISTO
        # -------------------------------------------------
        return JsonResponse({
            "ok": True,
            "msg": "Pago registrado y stock sumado.",
            "nuevo_stock": str(stock_obj.cantidad_disponible),
            "orden": oc.numero_orden,
        })

    except Exception as e:
        # 👇 esto es para que AHORA sí veas el error real en el popup
        return JsonResponse({"ok": False, "error": str(e)}, status=400)


@login_required
def paypal_ingresos_view(request):
    ordenes = (
        OrdenCompra.objects
        .filter(numero_orden__startswith="OC-PAYPAL-")
        .prefetch_related("lineas", "bodega", "proveedor")
        .order_by("-creado_en")[:50]
    )
    return render(request, "core/paypal_ingresos.html", {"ordenes": ordenes})

def _parse_contada_from_notas(notas: str | None):
    """
    Para RECUENTO_CIERRE: acepta 'contada=12.5' en notas.
    """
    if not notas:
        return None
    for part in notas.strip().split():
        if part.lower().startswith("contada="):
            try:
                return Decimal(part.split("=", 1)[1])
            except Exception:
                return None
    return None

@receiver(post_save, sender=MovimientoStock)
def rellenar_tablas_por_movimiento(sender, instance: MovimientoStock, created, **kwargs):
    """
    Se ejecuta SOLO cuando se CREA un MovimientoStock.
    - AJUSTE           -> crea AjusteInventario + LineaAjusteInventario
    - RECUENTO_CIERRE  -> crea RecuentoInventario + LineaRecuentoInventario (si hay 'contada=XX' en notas)
    - RESERVA          -> crea Reserva
    - LIBERAR_RESERVA  -> resta a la última Reserva coincidente (o por referencia)
    - TRANSFERENCIA / ENTRADA / SALIDA -> no hacen nada extra aquí (solo kárdex)
    IMPORTANTE: aquí NO se modifica la tabla Stock.
    """
    if not created:
        return

    codigo = (instance.tipo_movimiento.codigo or "").upper()

    with transaction.atomic():
        # ======================
        # AJUSTE
        # ======================
        if codigo == "AJUSTE":
            # Deducir bodega desde alguna ubicación
            bodega: Bodega | None = None
            for ub in (
                instance.ubicacion_bodega_hasta,
                instance.ubicacion_bodega_desde,
                instance.ubicacion_sucursal_hasta,
                instance.ubicacion_sucursal_desde,
            ):
                if ub is not None and hasattr(ub, "bodega"):
                    bodega = ub.bodega
                    break

            cab = AjusteInventario.objects.create(
                bodega=bodega,
                motivo=(instance.notas or "Ajuste inventario"),
                estado="CLOSED",
                creado_por=instance.creado_por,
            )
            LineaAjusteInventario.objects.create(
                ajuste=cab,
                producto=instance.producto,
                ubicacion_bodega=instance.ubicacion_bodega_hasta or instance.ubicacion_bodega_desde,
                ubicacion_sucursal=instance.ubicacion_sucursal_hasta or instance.ubicacion_sucursal_desde,
                lote=instance.lote,
                serie=instance.serie,
                cantidad_delta=Decimal(instance.cantidad),
                motivo=(instance.notas or ""),
            )
            return

        # ======================
        # RECUENTO_CIERRE
        # ======================
        if codigo == "RECUENTO_CIERRE":
            contada = _parse_contada_from_notas(instance.notas)
            if contada is None:
                return

            ub_bod = instance.ubicacion_bodega_hasta or instance.ubicacion_bodega_desde
            ub_suc = instance.ubicacion_sucursal_hasta or instance.ubicacion_sucursal_desde

            # Solo LECTURA para mostrar “cantidad_sistema” en la línea del recuento
            cant_sistema = Decimal("0")
            st = Stock.objects.filter(
                producto=instance.producto,
                ubicacion_bodega=ub_bod,
                ubicacion_sucursal=ub_suc,
            ).first()
            if st and st.cantidad_disponible is not None:
                cant_sistema = Decimal(st.cantidad_disponible)

            diferencia = Decimal(contada) - cant_sistema

            bodega: Bodega | None = None
            for ub in (ub_bod, ub_suc):
                if ub is not None and hasattr(ub, "bodega"):
                    bodega = ub.bodega
                    break

            rec = RecuentoInventario.objects.create(
                bodega=bodega,
                codigo_ciclo=timezone.now().strftime("C-%Y%m%d%H%M"),
                estado="CLOSED",
                creado_por=instance.creado_por,
            )
            LineaRecuentoInventario.objects.create(
                recuento=rec,
                producto=instance.producto,
                ubicacion_bodega=ub_bod,
                ubicacion_sucursal=ub_suc,
                lote=instance.lote,
                serie=instance.serie,
                cantidad_sistema=cant_sistema,
                cantidad_contada=contada,
                contado_por=instance.creado_por,
                diferencia=diferencia,
            )
            return

        # ======================
        # RESERVA
        # ======================
        if codigo == "RESERVA":
            Reserva.objects.create(
                producto=instance.producto,
                ubicacion_bodega=instance.ubicacion_bodega_hasta,
                ubicacion_sucursal=instance.ubicacion_sucursal_hasta,
                lote=instance.lote,
                serie=instance.serie,
                cantidad_reservada=Decimal(instance.cantidad),
                tabla_referencia=instance.tabla_referencia or "",
                referencia_id=instance.referencia_id,
            )
            return

        # ======================
        # LIBERAR_RESERVA
        # ======================
        if codigo == "LIBERAR_RESERVA":
            q = Reserva.objects.filter(
                producto=instance.producto,
                ubicacion_bodega=instance.ubicacion_bodega_hasta,
                ubicacion_sucursal=instance.ubicacion_sucursal_hasta,
            )
            if instance.referencia_id:
                q = q.filter(
                    tabla_referencia=instance.tabla_referencia or "",
                    referencia_id=instance.referencia_id,
                )
            r = q.order_by("-id").first()
            if r:
                nueva = (r.cantidad_reservada or Decimal("0")) - Decimal(instance.cantidad)
                r.cantidad_reservada = (nueva if nueva > 0 else Decimal("0"))
                r.save()
            return

        # TRANSFERENCIA / ENTRADA / SALIDA → no-ops aquí (solo guardamos el kárdex en tus vistas).
        return
    

@login_required
def auditoria_inventario(request):
    """
    Visual para inspeccionar lo que se rellena automáticamente desde MovimientoStock (por señales).
    Filtros:
      - ?desde=YYYY-MM-DD
      - ?hasta=YYYY-MM-DD
      - ?producto_id=123
    Por defecto: hoy.
    """
    hoy = timezone.localdate()
    desde_str = request.GET.get("desde")
    hasta_str = request.GET.get("hasta")
    producto_id = request.GET.get("producto_id")

    try:
        desde = datetime.fromisoformat(desde_str).date() if desde_str else hoy
    except Exception:
        desde = hoy
    try:
        hasta = datetime.fromisoformat(hasta_str).date() if hasta_str else hoy
    except Exception:
        hasta = hoy

    prod_filter = Q()
    if producto_id and producto_id != 'None':  # Verificar que producto_id no sea 'None'
        try:
            prod_filter = Q(producto_id=int(producto_id))
        except ValueError:
            prod_filter = Q()  # Si el valor no es un número válido, no agregar filtro

    # MOVS (solo por fecha, ignorando hora)
    movimientos = (
        MovimientoStock.objects
        .select_related("tipo_movimiento", "producto", "unidad",
                        "ubicacion_bodega_desde", "ubicacion_bodega_hasta",
                        "ubicacion_sucursal_desde", "ubicacion_sucursal_hasta",
                        "creado_por")
        .filter(ocurrido_en__date__range=(desde, hasta))
        .filter(prod_filter)
        .order_by("-ocurrido_en")[:500]
    )

    # AJUSTES
    ajustes = (
        AjusteInventario.objects
        .select_related("bodega", "creado_por")
        .filter(creado_en__date__range=(desde, hasta))
        .order_by("-creado_en")[:200]
    )
    lineas_ajuste = (
        LineaAjusteInventario.objects
        .select_related("ajuste", "producto", "ubicacion_bodega", "ubicacion_sucursal", "lote", "serie")
        .filter(ajuste__creado_en__date__range=(desde, hasta))
        .filter(prod_filter)
        .order_by("-ajuste__creado_en")[:1000]
    )

    # RECUENTOS
    recuentos = (
        RecuentoInventario.objects
        .select_related("bodega", "creado_por")
        .filter(creado_en__date__range=(desde, hasta))
        .order_by("-creado_en")[:200]
    )
    lineas_recuento = (
        LineaRecuentoInventario.objects
        .select_related("recuento", "producto", "ubicacion_bodega", "ubicacion_sucursal", "lote", "serie", "contado_por")
        .filter(recuento__creado_en__date__range=(desde, hasta))
        .filter(prod_filter)
        .order_by("-recuento__creado_en")[:1000]
    )

    # Paginación para 'movimientos'
    movimientos_paginator = Paginator(movimientos, 10)
    page_number = request.GET.get('page')
    movimientos_page = movimientos_paginator.get_page(page_number)

    # Paginación para 'ajustes'
    ajustes_paginator = Paginator(ajustes, 10)
    ajustes_page = ajustes_paginator.get_page(page_number)

    # Paginación para 'lineas_ajuste'
    lineas_ajuste_paginator = Paginator(lineas_ajuste, 10)
    lineas_ajuste_page = lineas_ajuste_paginator.get_page(page_number)

    # Paginación para 'recuentos'
    recuentos_paginator = Paginator(recuentos, 10)
    recuentos_page = recuentos_paginator.get_page(page_number)

    # Paginación para 'lineas_recuento'
    lineas_recuento_paginator = Paginator(lineas_recuento, 10)
    lineas_recuento_page = lineas_recuento_paginator.get_page(page_number)

    # RESERVAS
    reservas = (
        Reserva.objects
        .select_related("producto", "ubicacion_bodega", "ubicacion_sucursal", "lote", "serie")
        .filter(creado_en__date__range=(desde, hasta))
        .filter(prod_filter)
        .order_by("-creado_en")[:500]
    )

    productos = Producto.objects.only("id", "nombre").order_by("nombre")[:1000]

    ctx = {
        "desde": desde,
        "hasta": hasta,
        "movimientos": movimientos_page,
        "ajustes": ajustes_page,
        "lineas_ajuste": lineas_ajuste_page,
        "recuentos": recuentos_page,
        "lineas_recuento": lineas_recuento_page,
        "reservas": reservas,
        "productos": productos,
        "producto_id": int(producto_id) if producto_id and producto_id != 'None' else None,
    }
    return render(request, "core/auditoria_inventario.html", ctx)






def _bodega_from_stock(stock: Stock):
    """
    Retorna la bodega asociada al stock, ya sea:
    - Por su ubicación_bodega
    - O por la bodega que cuelga de la sucursal
    """
    if stock.ubicacion_bodega:
        return stock.ubicacion_bodega.bodega
    if stock.ubicacion_sucursal:
        suc = stock.ubicacion_sucursal.sucursal
        return suc.bodega if suc and suc.bodega else None
    return None


# ============================================================
#   PRE-SAVE: Guardar valor anterior
# ============================================================
@receiver(pre_save, sender=Stock)
def _record_prev_stock(sender, instance: Stock, **kwargs):
    """
    Guarda el valor anterior de cantidad_disponible antes del save()
    para poder comparar en el post_save.
    """
    if instance.pk:
        try:
            old = Stock.objects.get(pk=instance.pk)
            instance._prev_cantidad = old.cantidad_disponible or Decimal("0")
        except Stock.DoesNotExist:
            instance._prev_cantidad = Decimal("0")
    else:
        instance._prev_cantidad = Decimal("0")

from django.db.models.expressions import Expression
# ============================================================
#   POST-SAVE: Generar recuento cuando cambia el stock
# ============================================================
@receiver(post_save, sender=Stock)
def _crear_recuento_auto(sender, instance: Stock, created, **kwargs):
    """
    Cada vez que cambia la cantidad_disponible en Stock (suba o baje),
    crea automáticamente una línea en RecuentoInventario.

    IMPORTANTE: aquí nos aseguramos de NO usar expresiones F() ni similares,
    sino valores numéricos reales leídos desde la BD.
    """

    # ---- 1) Obtener el valor ACTUAL asegurándonos de que sea numérico ----
    actual_raw = instance.cantidad_disponible

    # Si viene una Expression (p.ej. F('cantidad_disponible')), recargamos desde BD
    if isinstance(actual_raw, Expression):
        actual = (
            Stock.objects
            .filter(pk=instance.pk)
            .values_list("cantidad_disponible", flat=True)
            .first()
            or Decimal("0")
        )
    else:
        actual = actual_raw or Decimal("0")

    # ---- 2) Obtener el valor PREVIO guardado en pre_save (o 0 si no hay) ----
    prev_raw = getattr(instance, "_prev_cantidad", None)

    if isinstance(prev_raw, Expression) or prev_raw is None:
        # Si por algún motivo es una Expression o no existe, asumimos 0
        prev = Decimal("0")
    else:
        prev = prev_raw or Decimal("0")

    # Solo si hay diferencia real
    if prev == actual:
        return

    # ---- 3) Determinar la bodega desde el stock ----
    bodega = _bodega_from_stock(instance)
    if not bodega:
        return  # no se puede asociar bodega ⇒ no se registra

    diff = actual - prev
    hoy = timezone.localdate()

    # ---- 4) Cabecera de recuento automática por día ----
    recuento, _ = RecuentoInventario.objects.get_or_create(
        bodega=bodega,
        codigo_ciclo=f"AUTO-{hoy.strftime('%Y%m%d')}",
        defaults={"estado": "OPEN"},
    )

    # ---- 5) Línea de detalle (solo valores numéricos, nada de F()) ----
    LineaRecuentoInventario.objects.create(
        recuento=recuento,
        producto=instance.producto,
        ubicacion_bodega=instance.ubicacion_bodega,
        ubicacion_sucursal=instance.ubicacion_sucursal,
        cantidad_sistema=prev,
        cantidad_contada=actual,
        diferencia=diff,
    )



# --- Finanzas: Vistas de reporte ---

# ---------- helpers ----------
def _auto_width(ws):
    """
    Ajusta el ancho de cada columna en una hoja de openpyxl
    según el texto más largo. Límite de ancho = 50.
    """
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            val = cell.value
            try:
                s = str(val) if val is not None else ""
            except Exception:
                s = ""
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

#----------------------------
# --- Finanzas: helpers comunes ---
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.db.models import Q, Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from django.contrib.auth.models import User
from io import BytesIO

from core.forms import FinanzasReporteForm
from core.models import (
    UsuarioPerfil, Bodega, Producto,
    OrdenCompra, RecepcionMercaderia, FacturaProveedor, Stock
)
def _is_auditor(user):
    """Permite AUDITOR, ADMIN y superuser."""
    try:
        if not user.is_authenticated:
            return False
        if getattr(user, "is_superuser", False):
            return True
        rol = getattr(getattr(user, "perfil", None), "rol", None)
        return rol in (UsuarioPerfil.Rol.AUDITOR, UsuarioPerfil.Rol.ADMIN)
    except Exception:
        return False

def _auto_width(ws):
    """Auto-anchos para openpyxl."""
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            val = cell.value
            s = str(val) if val is not None else ""
            max_len = max(max_len, len(s))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

def _build_finanzas_querysets(cleaned):
    """
    Devuelve (ordenes, recepciones, facturas, productos, resumen) aplicando los
    filtros del formulario ya limpiado.
    """
    bodega     = cleaned.get("bodega")
    proveedor  = cleaned.get("proveedor")
    fecha_desde = cleaned.get("fecha_desde")
    fecha_hasta = cleaned.get("fecha_hasta")

    # bases
    ordenes = OrdenCompra.objects.select_related("proveedor", "bodega").all()
    recepciones = RecepcionMercaderia.objects.select_related("bodega", "orden_compra").all()
    facturas = FacturaProveedor.objects.select_related("proveedor").all()
    productos = Producto.objects.all()

    # filtros por bodega
    if bodega:
        ordenes = ordenes.filter(bodega=bodega)
        recepciones = recepciones.filter(bodega=bodega)

        # productos con stock en esa bodega (ya sea en ubicación de bodega
        # o en sucursales que cuelgan de esa bodega)
        productos = (
            productos.filter(
                Q(stocks__ubicacion_bodega__bodega=bodega) |
                Q(stocks__ubicacion_sucursal__sucursal__bodega=bodega)
            )
            .annotate(
                stock_filtrado=Coalesce(
                    Sum(
                        "stocks__cantidad_disponible",
                        filter=Q(stocks__ubicacion_bodega__bodega=bodega) |
                                Q(stocks__ubicacion_sucursal__sucursal__bodega=bodega),
                    ),
                    Value(0, output_field=DecimalField(max_digits=20, decimal_places=6))
                )
            )
            .distinct()
        )
    else:
        # sin bodega, stock agregado general
        productos = productos.annotate(
            stock_filtrado=Coalesce(
                Sum("stocks__cantidad_disponible"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6))
            )
        )

    # filtros por proveedor
    if proveedor:
        ordenes = ordenes.filter(proveedor=proveedor)
        facturas = facturas.filter(proveedor=proveedor)
        # relación M2M Producto <-> Proveedor (ajusta related_name si varía)
        productos = productos.filter(usuarios_proveedor__proveedor=proveedor).distinct()

    # filtros por fecha
    if fecha_desde:
        ordenes = ordenes.filter(creado_en__date__gte=fecha_desde)
        recepciones = recepciones.filter(creado_en__date__gte=fecha_desde)
        facturas = facturas.filter(fecha_factura__gte=fecha_desde)
    if fecha_hasta:
        ordenes = ordenes.filter(creado_en__date__lte=fecha_hasta)
        recepciones = recepciones.filter(creado_en__date__lte=fecha_hasta)
        facturas = facturas.filter(fecha_factura__lte=fecha_hasta)

    resumen = {
        "ordenes_count": ordenes.count(),
        "recepciones_count": recepciones.count(),
        "facturas_count": facturas.count(),
        "productos_count": productos.count(),
        "stock_total": productos.aggregate(s=Sum("stock_filtrado"))["s"] or 0,
    }
    return (
        ordenes.order_by("-creado_en"),
        recepciones.order_by("-creado_en"),
        facturas.order_by("-fecha_factura"),
        productos.order_by("nombre"),
        resumen,
    )

# --- Página de reporte ---
@login_required
@user_passes_test(_is_auditor)
def finanzas_reporte(request):
    form = FinanzasReporteForm(request.GET or None)

    # poblar los combos
    form.fields["bodega"].queryset = Bodega.objects.all().order_by("codigo")
    form.fields["proveedor"].queryset = User.objects.filter(
        perfil__rol=UsuarioPerfil.Rol.PROVEEDOR, is_active=True
    ).order_by("username")

    # detectar si hay filtros aplicados
    has_filters = any([
        request.GET.get("bodega"),
        request.GET.get("proveedor"),
        request.GET.get("fecha_desde"),
        request.GET.get("fecha_hasta"),
    ])

    # si no hay filtros, no mostramos resultados
    if not has_filters:
        return render(request, "core/finanzas_reporte.html", {
            "form": form,
            "ordenes": [],
            "recepciones": [],
            "facturas": [],
            "productos": [],
            "has_filters": False,
            "resumen": {},
        })

    # si hay filtros, aplicamos búsqueda
    if form.is_valid():
        ordenes, recepciones, facturas, productos, resumen = _build_finanzas_querysets(form.cleaned_data)
    else:
        ordenes, recepciones, facturas, productos, resumen = [], [], [], [], {}

    return render(request, "core/finanzas_reporte.html", {
        "form": form,
        "ordenes": ordenes,
        "recepciones": recepciones,
        "facturas": facturas,
        "productos": productos,
        "has_filters": has_filters,
        "resumen": resumen,
    })

# --- Export: Excel (.xlsx) ---
@login_required
@user_passes_test(_is_auditor)
def finanzas_export_excel(request):
    form = FinanzasReporteForm(request.GET or None)
    if not form.is_valid():
        return HttpResponse("Filtros inválidos.", status=400)

    ordenes, recepciones, facturas, productos, resumen = _build_finanzas_querysets(form.cleaned_data)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
    except Exception:
        return HttpResponse("Falta instalar openpyxl (pip install openpyxl)", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Finanzas"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 1
    ws.cell(row=row, column=1, value="Reporte de Finanzas").font = Font(bold=True, size=14); row += 2

    # Órdenes
    ws.cell(row=row, column=1, value="Órdenes de compra").font = bold; row += 1
    headers = ["N° OC", "Proveedor", "Bodega", "Estado", "Fecha Esperada", "Creada"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h); cell.font = bold; cell.alignment = center; cell.border = border
    row += 1
    for oc in ordenes:
        ws.cell(row=row, column=1, value=oc.numero_orden).border = border
        ws.cell(row=row, column=2, value=(oc.proveedor.get_full_name() or oc.proveedor.username)).border = border
        ws.cell(row=row, column=3, value=f"{oc.bodega.codigo} - {oc.bodega.nombre}").border = border
        ws.cell(row=row, column=4, value=oc.estado).border = border
        ws.cell(row=row, column=5, value=(oc.fecha_esperada or "")).border = border
        ws.cell(row=row, column=6, value=oc.creado_en.strftime("%Y-%m-%d %H:%M")).border = border
        row += 1
    row += 1

    # Facturas
    ws.cell(row=row, column=1, value="Facturas proveedor").font = bold; row += 1
    headers = ["N°", "Proveedor", "Monto", "Fecha", "Estado"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h); cell.font = bold; cell.alignment = center; cell.border = border
    row += 1
    for f in facturas:
        ws.cell(row=row, column=1, value=f.numero_factura).border = border
        ws.cell(row=row, column=2, value=(f.proveedor.get_full_name() or f.proveedor.username)).border = border
        ws.cell(row=row, column=3, value=float(f.monto_total or 0)).border = border
        ws.cell(row=row, column=4, value=f.fecha_factura).border = border
        ws.cell(row=row, column=5, value=f.estado).border = border
        row += 1
    row += 1

    # Recepciones
    ws.cell(row=row, column=1, value="Recepciones").font = bold; row += 1
    headers = ["N°", "Bodega", "Estado", "Recibido en"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h); cell.font = bold; cell.alignment = center; cell.border = border
    row += 1
    for r in recepciones:
        ws.cell(row=row, column=1, value=r.numero_recepcion).border = border
        ws.cell(row=row, column=2, value=f"{r.bodega.codigo} - {r.bodega.nombre}").border = border
        ws.cell(row=row, column=3, value=r.estado).border = border
        ws.cell(row=row, column=4, value=(r.recibido_en.strftime("%Y-%m-%d %H:%M") if r.recibido_en else "")).border = border
        row += 1
    row += 1

    # Productos
    ws.cell(row=row, column=1, value="Productos (según filtros)").font = bold; row += 1
    headers = ["SKU", "Nombre", "Stock"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h); cell.font = bold; cell.alignment = center; cell.border = border
    row += 1
    for p in productos:
        ws.cell(row=row, column=1, value=p.sku).border = border
        ws.cell(row=row, column=2, value=p.nombre).border = border
        ws.cell(row=row, column=3, value=float(getattr(p, "stock_filtrado", 0) or 0)).border = border
        row += 1

    _auto_width(ws)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp['Content-Disposition'] = 'attachment; filename="reporte_finanzas.xlsx"'
    return resp

# --- Export: PDF (con tablas y columnas de productos más anchas) ---
@login_required
@user_passes_test(_is_auditor)
def finanzas_export_pdf(request):
    form = FinanzasReporteForm(request.GET or None)
    if not form.is_valid():
        return HttpResponse("Filtros inválidos.", status=400)

    ordenes, recepciones, facturas, productos, resumen = _build_finanzas_querysets(form.cleaned_data)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception:
        return HttpResponse("Falta instalar reportlab (pip install reportlab)", status=500)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []

    title = Paragraph("<b>Reporte de Finanzas</b>", styles["Title"])
    story += [title, Spacer(1, 10)]

    def section(title_txt, headers, rows, col_widths=None):
        story.append(Paragraph(f"<b>{title_txt}</b>", styles["Heading4"]))
        data = [headers] + (rows if rows else [[""] * len(headers)])
        tbl = Table(data, colWidths=col_widths)
        tbl.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 1, colors.black),
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("ALIGN", (0,0), (-1,0), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.extend([tbl, Spacer(1, 12)])

    # Órdenes
    section(
        "Órdenes de compra",
        ["N° OC", "Proveedor", "Bodega", "Estado", "Fecha Esperada", "Creada"],
        [[
            oc.numero_orden,
            (oc.proveedor.get_full_name() or oc.proveedor.username),
            f"{oc.bodega.codigo} - {oc.bodega.nombre}",
            oc.estado,
            (oc.fecha_esperada or ""),
            oc.creado_en.strftime("%Y-%m-%d %H:%M"),
        ] for oc in ordenes]
    )

    # Facturas
    section(
        "Facturas proveedor",
        ["N°", "Proveedor", "Monto", "Fecha", "Estado"],
        [[
            f.numero_factura,
            (f.proveedor.get_full_name() or f.proveedor.username),
            f"{f.monto_total}",
            f.fecha_factura.strftime("%Y-%m-%d"),
            f.estado,
        ] for f in facturas]
    )

    # Recepciones
    section(
        "Recepciones",
        ["N°", "Bodega", "Estado", "Recibido en"],
        [[
            r.numero_recepcion,
            f"{r.bodega.codigo} - {r.bodega.nombre}",
            r.estado,
            (r.recibido_en.strftime("%Y-%m-%d %H:%M") if r.recibido_en else ""),
        ] for r in recepciones]
    )

    # Productos (SKU y Nombre con más ancho para que no se solapen)
    section(
        "Productos (según filtros)",
        ["SKU", "Nombre", "Stock"],
        [[
            p.sku,
            p.nombre,
            f"{getattr(p, 'stock_filtrado', 0) or 0}",
        ] for p in productos],
        col_widths=[120, 320, 60],  # ← ajustado para evitar solapes
    )

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    resp = HttpResponse(pdf, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="reporte_finanzas.pdf"'
    return resp

# --- Cambio de moneda en sesión ---
def set_currency(request):
    cur = (request.GET.get("c") or "").upper()
    if cur in ("CLP", "USD"):
        request.session["currency"] = cur
    next_url = request.GET.get("next") or request.META.get("HTTP_REFERER") or reverse("products")
    return HttpResponseRedirect(next_url)
































@login_required
def bodega_agregar_sucursal(request, bodega_id):
    bodega = get_object_or_404(Bodega, pk=bodega_id)

    # traemos TODAS las sucursales para mostrarlas
    sucursales = (
        Sucursal.objects
        .select_related("bodega")
        .order_by("nombre")
    )

    if request.method == "POST":
        # puede venir 1 o muchas
        seleccionadas = request.POST.getlist("sucursales")
        if not seleccionadas:
            return render(
                request,
                "core/Bodega/bodega_agregar_sucursal.html",
                {
                    "bodega": bodega,
                    "sucursales": sucursales,
                    "error": "Debes seleccionar al menos una sucursal.",
                },
            )

        # reasignamos todas las seleccionadas a ESTA bodega
        Sucursal.objects.filter(pk__in=seleccionadas).update(bodega=bodega)

        return redirect("bodega-list")

    return render(
        request,
        "core/bodega_agregar_sucursal.html",
        {
            "bodega": bodega,
            "sucursales": sucursales,
        },
    )





@login_required
def bodega_productos(request, bodega_id: int):
    bodega = get_object_or_404(Bodega, pk=bodega_id)
    q = (request.GET.get("q") or "").strip()

    # Obtener el código de la bodega
    bodega_codigo = bodega.codigo

    # Construir el código de la ubicación predeterminada de la bodega
    default_location_code = f"{bodega_codigo}-000-000"

    # Ubicaciones activas de ESTA bodega, EXCLUYENDO la ubicación predeterminada
    # (estas son las que se muestran en el <select> del modal)
    ubicaciones = (
        UbicacionBodega.objects
        .filter(bodega=bodega, activo=True)
        .exclude(codigo=default_location_code)
        .order_by("codigo")
    )

    # ==== Productos con stock en la bodega (cualquier ubicación activa, incluida la default) ====
    productos = (
        Producto.objects
        .filter(
            stocks__ubicacion_bodega__bodega=bodega,
            stocks__ubicacion_bodega__activo=True,
        )
        .annotate(
            # Stock total disponible en la bodega (todas las ubicaciones activas de esa bodega,
            # incluyendo la ubicación default de recepción)
            stock_total=Coalesce(
                Sum(
                    "stocks__cantidad_disponible",
                    filter=Q(
                        stocks__ubicacion_bodega__bodega=bodega,
                        stocks__ubicacion_bodega__activo=True,
                    ),
                ),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            ),
            # Stock validado para movimiento (suma de cantidad_validada en esta bodega)
            stock_validado=Coalesce(
                Sum(
                    "stocks__cantidad_validada",
                    filter=Q(
                        stocks__ubicacion_bodega__bodega=bodega,
                        stocks__ubicacion_bodega__activo=True,
                    ),
                ),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            ),
        )
        .distinct()
        .order_by("nombre")
    )

    if q:
        productos = productos.filter(
            Q(nombre__icontains=q) | Q(sku__icontains=q)
        )

    total_items = productos.count()

    # Suma de TODOS los stock_total de los productos listados
    suma_stock = (
        productos.aggregate(
            s=Coalesce(
                Sum("stock_total"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            )
        )["s"] or 0
    )

    return render(
        request,
        "core/Movimientos/bodega_productos.html",
        {
            "bodega": bodega,
            "productos": productos,
            "ubicaciones": ubicaciones,
            "q": q,
            "total_items": total_items,
            "suma_stock": suma_stock,
        },
    )


@login_required
def ajax_ubicaciones_por_producto(request):

    bodega_id = request.GET.get("bodega_id")
    producto_id = request.GET.get("producto_id")

    if not bodega_id or not producto_id:
        return JsonResponse(
            {"ok": False, "error": "Faltan parámetros."},
            status=400,
        )

    try:
        qs = (
            Stock.objects
            .select_related("ubicacion_bodega")
            .filter(
                producto_id=producto_id,
                ubicacion_bodega__bodega_id=bodega_id,
            )
        )
    except Exception as e:
        return JsonResponse(
            {"ok": False, "error": f"Error en la consulta: {e}"},
            status=500,
        )

    ubicaciones = []
    for st in qs:
        if st.ubicacion_bodega is None:
            continue
        ubi = st.ubicacion_bodega
        ubicaciones.append(
            {
                "codigo": ubi.codigo,
                # usamos el campo `area` como descripción
                "nombre": ubi.area or "",
                "cantidad": float(st.cantidad_disponible),
            }
        )

    return JsonResponse(
        {
            "ok": True,
            "ubicaciones": ubicaciones,
        }
    )








@login_required
def paypal_ingresos_view(request):
    ordenes = (
        OrdenCompra.objects
        .filter(numero_orden__startswith="OC-PAYPAL-")
        .select_related("bodega")
        .prefetch_related("lineas__producto")
        .order_by("-id")
    )
    return render(request, "core/paypal_ingresos.html", {"ordenes": ordenes})
from django.contrib.auth.views import PasswordResetView





def guia_transferencia_detalle(request, pk):
    queryset = (
        Transferencia.objects
        .select_related(
            "bodega_origen",
            "bodega_destino",   # destino bodega (si existe)
            "sucursal_destino",
            "creado_por",
        )
        .prefetch_related(
            "lineas__producto",
            "lineas__unidad",
            "lineas__lote",
            "lineas__serie",
        )
    )

    transferencia = get_object_or_404(queryset, pk=pk)

    # Total de unidades (suma de todas las cantidades de las líneas)
    total_cantidad = (
        transferencia.lineas.aggregate(
            total=Coalesce(
                Sum("cantidad"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            )
        )["total"]
        or 0
    )

    return render(
        request,
        "core/guias/guia_transferencia.html",
        {
            "transferencia": transferencia,
            "total_cantidad": total_cantidad,
        },
    )




from django.core.paginator import Paginator
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
from .models import Transferencia  










































































































































































































































@login_required
def resumen_guias_despacho(request):
    """
    Listado de guías de despacho / transferencias,
    con filtro por tipo de movimiento.
    """
    tipo_raw = (request.GET.get("tipo", "") or "").strip()

    if tipo_raw == "":
        tipo = "TODOS"
    elif tipo_raw == "bod_bod":
        tipo = "BOD_BOD"
    elif tipo_raw == "con_sucursal":
        tipo = "BOD_SUC"
    elif tipo_raw == "suc_bod":
        tipo = "SUC_BOD"
    elif tipo_raw == "suc_suc":
        tipo = "SUC_SUC"
    else:
        tipo = tipo_raw.upper()

    # Base: todas las transferencias con relaciones necesarias
    base_qs = (
        Transferencia.objects
        .select_related(
            "bodega_origen",
            "bodega_destino",
            "sucursal_origen",
            "sucursal_destino",
            "creado_por",
        )
        .order_by("-creado_en")
    )

    # Filtrar según tipo de movimiento
    if tipo == "TODOS":
        transferencias = base_qs
    elif tipo == "BOD_BOD":
        transferencias = base_qs.filter(tipo_movimiento="BOD_BOD")
    elif tipo == "BOD_SUC":
        transferencias = base_qs.filter(tipo_movimiento="BOD_SUC")
    elif tipo == "SUC_BOD":
        transferencias = base_qs.filter(tipo_movimiento="SUC_BOD")
    elif tipo == "SUC_SUC":
        transferencias = base_qs.filter(tipo_movimiento="SUC_SUC")
    else:
        transferencias = base_qs

    # Paginación: 20 transferencias por página
    paginator = Paginator(transferencias, 13)  # 20 por página
    page_number = request.GET.get("page", 1)

    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'transferencias': page_obj,
        'tipo_sel': tipo,
    }
    return render(request, "core/Guias/resumen_guias.html", context)

















def agregar_ubicacion_bodega(request, bodega_id):
    bodega = get_object_or_404(Bodega, pk=bodega_id)

    if request.method == "POST":
        form = UbicacionBodegaForm(request.POST)
        if form.is_valid():
            ubicacion = form.save(commit=False)
            ubicacion.bodega = bodega  # ahora sí tiene bodega

            area_cod = form.cleaned_data["area_codigo"]
            estante_cod = form.cleaned_data["estante_codigo"]
            ubicacion.set_codigo(area_cod, estante_cod)

            ubicacion.save()
            return redirect("bodega-list")
    else:
        form = UbicacionBodegaForm()

    return render(
        request,
        "core/Movimientos/agregar.html",
        {"form": form, "bodega": bodega},
    )




def mover_producto_ubicacion(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == "POST":
        nueva_ubicacion_id = request.POST.get("ubicacion_id")
        nueva_ubicacion = get_object_or_404(UbicacionBodega, id=nueva_ubicacion_id)
        producto.ubicacion = nueva_ubicacion
        producto.save()
        return HttpResponse("Producto movido correctamente")
    
    ubicaciones = UbicacionBodega.objects.all()
    return render(request, "core/Movimientos/mover_producto.html", {"producto": producto, "ubicaciones": ubicaciones})





# Configuración de logs
logger = logging.getLogger(__name__)


@login_required
@require_http_methods(["GET", "POST"])
def validar_stock_bodega(request, bodega_id: int):
    bodega = get_object_or_404(Bodega, pk=bodega_id)

    # ---------- POST: guardar validación ----------
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except json.JSONDecodeError:
            return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

        lineas = payload.get("lineas", [])
        if not isinstance(lineas, list):
            return JsonResponse({"ok": False, "error": "Formato de datos incorrecto."}, status=400)

        try:
            with transaction.atomic():
                for linea in lineas:
                    stock_id = linea.get("stock_id")
                    raw_val = linea.get("validada", "0")

                    try:
                        cant_validada = Decimal(str(raw_val or "0"))
                    except Exception:
                        return JsonResponse(
                            {"ok": False, "error": "Cantidad inválida."},
                            status=400,
                        )

                    # Buscar el stock en la bodega
                    stock = get_object_or_404(
                        Stock,
                        pk=stock_id,
                        ubicacion_bodega__bodega=bodega,
                    )

                    # Validaciones simples
                    if cant_validada < 0:
                        return JsonResponse(
                            {"ok": False, "error": "La cantidad no puede ser negativa."},
                            status=400,
                        )

                    # Validar que la cantidad no sea mayor que la disponible
                    if cant_validada > stock.cantidad_disponible:
                        return JsonResponse(
                            {
                                "ok": False,
                                "error": f"No hay suficiente stock para {stock.producto.sku}. Stock disponible: {stock.cantidad_disponible}.",
                            },
                            status=400,
                        )

                    # Si la validación pasa, actualizamos el stock
                    stock.cantidad_validada = cant_validada
                    stock.save(update_fields=["cantidad_validada"])

        except Exception as e:
            # Por si hay cualquier error de BD
            return JsonResponse(
                {"ok": False, "error": f"Error al guardar la validación: {e}"},
                status=500,
            )

        return JsonResponse(
            {"ok": True, "message": "Validación guardada correctamente."}
        )

    # ---------- GET: mostrar pantalla ----------
    q = (request.GET.get("q") or "").strip()
    producto_id = request.GET.get("producto")  # Obtener el ID del producto de la URL

    # Obtener el código de la bodega (ahora lo tomamos desde el campo 'codigo')
    bodega_codigo = bodega.codigo

    # Construir el código de la ubicación default para esta bodega
    default_location_code = f"{bodega_codigo}-000-000"

    # Filtrar productos por la bodega y excluir la ubicación default
    filas = (
        Stock.objects.filter(ubicacion_bodega__bodega=bodega)
        .select_related("producto", "ubicacion_bodega")
        .exclude(ubicacion_bodega__codigo=default_location_code)  # Excluir la ubicación de default
        .exclude(ubicacion_bodega__codigo__in=["OTRA-UBICACION"])  # Excluir otras ubicaciones no válidas
        .order_by("producto__sku", "ubicacion_bodega__codigo")
    )

    if producto_id:
        filas = filas.filter(producto_id=producto_id)  # Filtro por producto

    if q:
        filas = filas.filter(
            Q(producto__sku__icontains=q)
            | Q(producto__nombre__icontains=q)
            | Q(ubicacion_bodega__codigo__icontains=q)
        )

    agg = filas.aggregate(
        total_disponible=Coalesce(
            Sum("cantidad_disponible"),
            Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
        ),
        total_validado=Coalesce(
            Sum("cantidad_validada"),
            Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
        ),
    )

    context = {
        "bodega": bodega,
        "filas": filas,
        "total_disponible": agg["total_disponible"] or 0,
        "total_validado": agg["total_validado"] or 0,
        "q": q,
    }
    return render(request, "core/Movimientos/validar_stock_bodega.html", context)

@login_required
@require_POST
def ajax_validar_stock_ubicacion(request):
    """
    Recibe: bodega_id, producto_id, ubicacion_id, cantidad_validada
    y guarda el cupo validado en Stock.cantidad_validada.
    """
    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

    bodega_id     = data.get("bodega_id")
    producto_id   = data.get("producto_id")
    ubicacion_id  = data.get("ubicacion_id")
    cant_str      = data.get("cantidad")

    if not (bodega_id and producto_id and ubicacion_id):
        return JsonResponse({"ok": False, "error": "Faltan parámetros."}, status=400)

    try:
        cantidad = Decimal(str(cant_str or "0"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Cantidad inválida."}, status=400)

    # Buscar el stock en ESA ubicación de ESA bodega
    try:
        stock = (
            Stock.objects
            .select_related("ubicacion_bodega")
            .get(
                producto_id=producto_id,
                ubicacion_bodega_id=ubicacion_id,
                ubicacion_bodega__bodega_id=bodega_id,
            )
        )
    except Stock.DoesNotExist:
        return JsonResponse({"ok": False, "error": "No existe stock en esa ubicación."}, status=404)

    # Validar que la cantidad no sea negativa
    if cantidad < 0:
        return JsonResponse({"ok": False, "error": "La cantidad no puede ser negativa."}, status=400)

    # Validar que la cantidad no supere el stock disponible
    if cantidad > stock.cantidad_disponible:
        return JsonResponse(
            {
                "ok": False,
                "error": f"No puedes validar más de lo disponible ({stock.cantidad_disponible}).",
            },
            status=400,
        )

    # Actualizar la cantidad validada en el stock
    stock.cantidad_validada = cantidad
    stock.save(update_fields=["cantidad_validada"])

    return JsonResponse(
        {
            "ok": True,
            "message": f"Se validaron {cantidad} unidades para movimientos.",
            "cantidad_validada": float(stock.cantidad_validada),
        }
    )








##esta es la funcion de ubicacion
@login_required
@require_POST
def ajax_asignar_producto_ubicacion(request):
    """
    Asigna stock de un producto a una ubicación de la bodega.

    Prioridad para descontar stock:
      1) Ubicación 'SIN UBICACIÓN' (default).
      2) Otras ubicaciones del mismo producto en la bodega (excepto destino).

    Espera JSON:
      {
        "bodega_id": 1,
        "producto_id": 10,
        "ubicacion_id": 5,   # destino
        "cantidad": 3
      }
    """
    # ---------- Parsear JSON ----------
    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

    bodega_id = data.get("bodega_id")
    producto_id = data.get("producto_id")
    ubicacion_id = data.get("ubicacion_id")
    cantidad = data.get("cantidad", 0)

    if not (producto_id and ubicacion_id and bodega_id):
        return JsonResponse({"ok": False, "error": "Datos incompletos"}, status=400)

    # ---------- Buscar producto y ubicación ----------
    try:
        producto = Producto.objects.get(pk=producto_id)
        ubicacion = UbicacionBodega.objects.get(pk=ubicacion_id, bodega_id=bodega_id)
    except (Producto.DoesNotExist, UbicacionBodega.DoesNotExist):
        return JsonResponse({"ok": False, "error": "Producto o ubicación no encontrados"}, status=404)

    # ---------- Normalizar cantidad ----------
    try:
        cantidad_val = Decimal(str(cantidad))
        if cantidad_val < 0:
            cantidad_val = Decimal("0")
    except Exception:
        cantidad_val = Decimal("0")

    # ---------- Verificar stock total ANTES de mover ----------
    total_disp = (
        Stock.objects
        .filter(producto=producto, ubicacion_bodega__bodega_id=bodega_id)
        .aggregate(
            t=Coalesce(
                Sum("cantidad_disponible"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            )
        )["t"]
        or Decimal("0")
    )

    # Asegurarse de que cantidad y total_disp sean ambos de tipo Decimal antes de la comparación
    if cantidad_val > total_disp:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    f"No hay stock suficiente del producto {producto.sku} "
                    f"en la bodega para mover {cantidad_val} unidades. "
                    f"Disponible total: {total_disp}."
                ),
            },
            status=400,
        )

    # ---------- Ubicación 'SIN UBICACIÓN' (default) de la bodega ----------
    ubi_default = UbicacionBodega.objects.filter(bodega_id=bodega_id, codigo="SIN UBICACIÓN").first()

    # ---------- Movimiento de stock con prioridad ----------
    with transaction.atomic():
        # 1) stock en SIN UBICACIÓN (puede no existir)
        stock_default = (
            Stock.objects
            .select_for_update()
            .filter(producto=producto, ubicacion_bodega=ubi_default)
            .first()
        )

        # 2) otras ubicaciones (fuentes secundarias), EXCEPTO destino y default
        otros_stocks = (
            Stock.objects
            .select_for_update()
            .filter(
                producto=producto,
                ubicacion_bodega__bodega_id=bodega_id,
            )
            .exclude(ubicacion_bodega=ubicacion)
            .exclude(ubicacion_bodega=ubi_default)
            .order_by("ubicacion_bodega__codigo")
        )

        # 3) destino
        stock_destino, _ = (
            Stock.objects
            .select_for_update()
            .get_or_create(
                producto=producto,
                ubicacion_bodega=ubicacion,
                defaults={"cantidad_disponible": Decimal("0")},
            )
        )

        falta = cantidad_val

        # ----- Prioridad 1: SIN UBICACIÓN -----
        if stock_default and stock_default.cantidad_disponible > 0 and falta > 0:
            mover = min(stock_default.cantidad_disponible, falta)
            stock_default.cantidad_disponible -= mover
            stock_destino.cantidad_disponible += mover
            falta -= mover
            stock_default.save(update_fields=["cantidad_disponible"])

        # ----- Prioridad 2: otras ubicaciones -----
        for s in otros_stocks:
            if falta <= 0:
                break

            if s.cantidad_disponible <= 0:
                continue

            mover = min(s.cantidad_disponible, falta)
            s.cantidad_disponible -= mover
            stock_destino.cantidad_disponible += mover
            falta -= mover
            s.save(update_fields=["cantidad_disponible"])

        # Por seguridad: si aquí aún falta algo, lanzamos excepción para forzar rollback
        if falta > 0:
            raise RuntimeError(
                f"Inconsistencia: falta {falta} unidades al mover stock de {producto.sku}."
            )

        stock_destino.save(update_fields=["cantidad_disponible"])

    return JsonResponse({
        "ok": True,
        "message": (
            f"Se movieron {cantidad_val} unidades hacia {ubicacion.codigo}, "
            f"tomando stock disponible de la bodega (prioridad 'SIN UBICACIÓN')."
        ),
    })



























































































































































@login_required
@require_POST
def ajax_asignar_producto_ubicacion_sucursal(request):
    """
    Asigna stock de un producto a una ubicación interna de la SUCURSAL.

    Prioridad para descontar stock:
      1) Ubicación 'SIN UBICACIÓN' (default).
      2) Otras ubicaciones del mismo producto en la sucursal (excepto destino).

    Espera JSON:
      {
        "sucursal_id": 1,
        "producto_id": 10,
        "ubicacion_id": 5,   # destino
        "cantidad": 3
      }
    """
    # ---------- Parsear JSON ----------
    try:
        data = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

    sucursal_id  = data.get("sucursal_id")
    producto_id  = data.get("producto_id")
    destino_id   = data.get("ubicacion_id")
    cant_raw     = data.get("cantidad")

    if not sucursal_id or not producto_id or not destino_id:
        return JsonResponse(
            {"ok": False, "error": "Faltan datos obligatorios (sucursal, producto o ubicación destino)."},
            status=400,
        )

    # ---------- Objetos base ----------
    sucursal = get_object_or_404(Sucursal, pk=sucursal_id)
    producto = get_object_or_404(Producto, pk=producto_id)
    ubi_destino = get_object_or_404(
        UbicacionSucursal,
        pk=destino_id,
        sucursal=sucursal,
    )

    # ---------- Validar cantidad ----------
    try:
        cantidad = Decimal(str(cant_raw or "0"))
    except Exception:
        return JsonResponse(
            {"ok": False, "error": "Cantidad inválida."},
            status=400,
        )

    if cantidad <= 0:
        return JsonResponse(
            {"ok": False, "error": "La cantidad debe ser mayor a 0."},
            status=400,
        )

    # ---------- Verificar stock total ANTES de mover ----------
    total_disp = (
        Stock.objects
        .filter(producto=producto, ubicacion_sucursal__sucursal=sucursal)
        .aggregate(
            t=Coalesce(
                Sum("cantidad_disponible"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            )
        )["t"]
        or Decimal("0")
    )

    if cantidad > total_disp:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    f"No hay stock suficiente del producto {producto.sku} "
                    f"en la sucursal para mover {cantidad} unidades. "
                    f"Disponible total: {total_disp}."
                ),
            },
            status=400,
        )

    # Ubicación 'SIN UBICACIÓN' (default) de la sucursal
    ubi_default = ensure_ubicacion_sucursal(sucursal)

    # ---------- Movimiento de stock con prioridad ----------
    with transaction.atomic():
        # 1) stock en SIN UBICACIÓN (puede no existir)
        stock_default = (
            Stock.objects
            .select_for_update()
            .filter(producto=producto, ubicacion_sucursal=ubi_default)
            .first()
        )

        # 2) otras ubicaciones (fuentes secundarias), EXCEPTO destino y default
        otros_stocks = (
            Stock.objects
            .select_for_update()
            .filter(
                producto=producto,
                ubicacion_sucursal__sucursal=sucursal,
            )
            .exclude(ubicacion_sucursal=ubi_destino)
            .exclude(ubicacion_sucursal=ubi_default)
            .order_by("ubicacion_sucursal__codigo")
        )

        # 3) destino
        stock_destino, _ = (
            Stock.objects
            .select_for_update()
            .get_or_create(
                producto=producto,
                ubicacion_sucursal=ubi_destino,
                defaults={"cantidad_disponible": Decimal("0")},
            )
        )

        falta = cantidad

        # ----- Prioridad 1: SIN UBICACIÓN -----
        if stock_default and stock_default.cantidad_disponible > 0 and falta > 0:
            mover = min(stock_default.cantidad_disponible, falta)
            stock_default.cantidad_disponible -= mover
            stock_destino.cantidad_disponible += mover
            falta -= mover
            stock_default.save(update_fields=["cantidad_disponible"])

        # ----- Prioridad 2: otras ubicaciones -----
        for s in otros_stocks:
            if falta <= 0:
                break

            if s.cantidad_disponible <= 0:
                continue

            mover = min(s.cantidad_disponible, falta)
            s.cantidad_disponible -= mover
            stock_destino.cantidad_disponible += mover
            falta -= mover
            s.save(update_fields=["cantidad_disponible"])

        # Por seguridad: si aquí aún falta algo, lanzamos excepción para forzar rollback
        if falta > 0:
            raise RuntimeError(
                f"Inconsistencia: falta {falta} unidades al mover stock de {producto.sku}."
            )

        stock_destino.save(update_fields=["cantidad_disponible"])

    return JsonResponse({
        "ok": True,
        "message": (
            f"Se movieron {cantidad} unidades hacia {ubi_destino.codigo}, "
            f"tomando stock disponible de la sucursal (prioridad 'SIN UBICACIÓN')."
        ),
    })


@login_required
@require_POST
def ajax_validar_stock_ubicacion_sucursal(request):
    """
    Valida stock para una ubicación específica de SUCURSAL.
    """
    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

    sucursal_id  = data.get("sucursal_id")
    producto_id  = data.get("producto_id")
    ubicacion_id = data.get("ubicacion_id")
    cant_str     = data.get("cantidad")

    if not (sucursal_id and producto_id and ubicacion_id):
        return JsonResponse({"ok": False, "error": "Faltan parámetros."}, status=400)

    try:
        cantidad = Decimal(str(cant_str or "0"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Cantidad inválida."}, status=400)

    try:
        stock = (
            Stock.objects
            .select_related("ubicacion_sucursal")
            .get(
                producto_id=producto_id,
                ubicacion_sucursal_id=ubicacion_id,
                ubicacion_sucursal__sucursal_id=sucursal_id,
            )
        )
    except Stock.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Stock inexistente en esa ubicación."}, status=404)

    if cantidad < 0:
        return JsonResponse({"ok": False, "error": "No puede ser negativo."}, status=400)

    if cantidad > stock.cantidad_disponible:
        return JsonResponse({
            "ok": False,
            "error": f"No puedes validar más de lo disponible ({stock.cantidad_disponible})."
        }, status=400)

    stock.cantidad_validada = cantidad
    stock.save(update_fields=["cantidad_validada"])

    return JsonResponse({
        "ok": True,
        "message": f"Validado {cantidad} unidades.",
        "cantidad_validada": float(stock.cantidad_validada),
    })


@login_required
def ajax_ubicaciones_sucursal_por_producto(request):
    """
    Devuelve todas las ubicaciones (default + internas) del PRODUCTO en la SUCURSAL.
    Se usa para cargar las “pills” dentro de la tabla.
    """
    sucursal_id = request.GET.get("sucursal_id")
    producto_id = request.GET.get("producto_id")

    if not (sucursal_id and producto_id):
        return JsonResponse({"ok": False, "error": "Parámetros incompletos."}, status=400)

    try:
        sucursal = Sucursal.objects.get(pk=sucursal_id)
    except Sucursal.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Sucursal no existe."}, status=404)

    ubicaciones_qs = (
        Stock.objects
        .select_related("ubicacion_sucursal")
        .filter(
            producto_id=producto_id,
            ubicacion_sucursal__sucursal_id=sucursal_id
        )
        .values(
            "ubicacion_sucursal_id",
            "ubicacion_sucursal__codigo",
            "ubicacion_sucursal__area",   # tu campo real
            "cantidad_disponible",
            "cantidad_validada",
        )
    )

    ubicaciones = []
    for u in ubicaciones_qs:
        ubicaciones.append({
            "id": u["ubicacion_sucursal_id"],
            "codigo": u["ubicacion_sucursal__codigo"],
            "area": u["ubicacion_sucursal__area"],
            "cantidad": float(u["cantidad_disponible"]),
            "validada": float(u["cantidad_validada"]) if u["cantidad_validada"] else 0
        })

    return JsonResponse({
        "ok": True,
        "ubicaciones": ubicaciones,
    })


@login_required
def ver_stock_sucursal(request, sucursal_id):
    """
    Versión 'Productos de bodega', pero para SUCURSAL.
    Lista productos que tienen stock en cualquier ubicación de esa sucursal.
    """
    sucursal = get_object_or_404(Sucursal, pk=sucursal_id)

    # (opcional) asegurar que exista al menos una ubicación para la sucursal
    try:
        ensure_ubicacion_sucursal(sucursal)
    except Exception:
        pass

    # texto de búsqueda (buscar por SKU o nombre)
    q = (request.GET.get("q") or "").strip()

    # Obtener el ID del producto de la URL, si se pasa
    producto_id = request.GET.get("producto")

    # === Productos con stock en la sucursal ===
    productos_qs = (
        Producto.objects.filter(
            stocks__ubicacion_sucursal__sucursal=sucursal
        )
        .annotate(
            stock_sucursal=Coalesce(
                Sum(
                    "stocks__cantidad_disponible",
                    filter=Q(stocks__ubicacion_sucursal__sucursal=sucursal),
                ),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            ),
            stock_validado=Coalesce(
                Sum(
                    "stocks__cantidad_validada",
                    filter=Q(stocks__ubicacion_sucursal__sucursal=sucursal),
                ),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            ),
        )
        .filter(stock_sucursal__gt=0)  # solo productos con stock en la sucursal
        .exclude(stocks__ubicacion_sucursal__codigo__in=["BOD-01-000-000", "OTRA-UBICACION"])  # Excluir ubicaciones no válidas
        .distinct()
        .order_by("sku")
    )

    # Si se pasa un producto en la URL, filtrar por ese producto
    if producto_id:
        productos_qs = productos_qs.filter(id=producto_id)

    # Filtro por SKU / nombre
    if q:
        productos_qs = productos_qs.filter(
            Q(sku__icontains=q) | Q(nombre__icontains=q)
        )

    # Métricas de productos
    total_items = productos_qs.count()
    suma_stock = (
        productos_qs.aggregate(
            t=Coalesce(
                Sum("stock_sucursal"),
                Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
            )
        )["t"]
        or 0
    )

    # Excluir la ubicación predeterminada de la sucursal
    ubicaciones = (
        UbicacionSucursal.objects.filter(sucursal=sucursal, activo=True)
        .exclude(codigo=f"{sucursal.codigo}-000-000")  # Excluir ubicación predeterminada
        .order_by("codigo")
    )

    # Pasar solo la cantidad al contexto para mostrar en el template
    context = {
        "sucursal": sucursal,
        "productos": productos_qs,
        "ubicaciones": ubicaciones,
        "total_items": total_items,
        "suma_stock": suma_stock,
        "q": q,
    }

    return render(request, "core/Movimientos/sucursal_productos.html", context)



@login_required
@require_http_methods(["GET", "POST"])
def validar_stock_sucursal(request, sucursal_id):
    sucursal = get_object_or_404(Sucursal, pk=sucursal_id)

    # ---------- POST: guardar validación ----------
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except json.JSONDecodeError:
            return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

        lineas = payload.get("lineas", [])
        if not isinstance(lineas, list):
            return JsonResponse({"ok": False, "error": "Formato de datos incorrecto."}, status=400)

        try:
            with transaction.atomic():
                for linea in lineas:
                    stock_id = linea.get("stock_id")
                    raw_val = linea.get("validada", "0")

                    try:
                        cant_validada = Decimal(str(raw_val or "0"))
                    except Exception:
                        return JsonResponse(
                            {"ok": False, "error": "Cantidad inválida."},
                            status=400,
                        )

                    stock = get_object_or_404(
                        Stock,
                        pk=stock_id,
                        ubicacion_sucursal__sucursal=sucursal,
                    )

                    # Validaciones simples
                    if cant_validada < 0:
                        return JsonResponse(
                            {"ok": False, "error": "La cantidad no puede ser negativa."},
                            status=400,
                        )
                    if cant_validada > stock.cantidad_disponible:
                        return JsonResponse(
                            {
                                "ok": False,
                                "error": f"Cantidad mayor al disponible para {stock.producto.sku}.",
                            },
                            status=400,
                        )

                    stock.cantidad_validada = cant_validada
                    stock.save(update_fields=["cantidad_validada"])

        except Exception as e:
            # Por si hay cualquier error de BD
            return JsonResponse(
                {"ok": False, "error": f"Error al guardar la validación: {e}"},
                status=500,
            )

        return JsonResponse(
            {"ok": True, "message": "Validación guardada correctamente."}
        )

    # ---------- GET: mostrar pantalla ----------
    q = (request.GET.get("q") or "").strip()

    # Obtener el ID del producto de la URL, si se pasa
    producto_id = request.GET.get("producto")

    # Obtener el código de la sucursal (ahora lo tomamos desde el campo 'codigo')
    sucursal_codigo = sucursal.codigo

    # Construir el código de la ubicación default para esta sucursal
    default_location_code = f"{sucursal_codigo}-000-000"

    # === Productos con stock en la sucursal ===
    filas = (
        Stock.objects.filter(ubicacion_sucursal__sucursal=sucursal)
        .select_related("producto", "ubicacion_sucursal")
        .exclude(ubicacion_sucursal__codigo=default_location_code)  # Excluir la ubicación de default
        .exclude(ubicacion_sucursal__codigo__in=["BOD-01-000-000", "OTRA-UBICACION"])  # Excluir otras ubicaciones no válidas
        .order_by("producto__sku", "ubicacion_sucursal__codigo")
    )

    # Si se pasa un producto en la URL, filtrar por ese producto
    if producto_id:
        filas = filas.filter(producto_id=producto_id)

    if q:
        filas = filas.filter(
            Q(producto__sku__icontains=q)
            | Q(producto__nombre__icontains=q)
            | Q(ubicacion_sucursal__codigo__icontains=q)
        )

    agg = filas.aggregate(
        total_disponible=Coalesce(
            Sum("cantidad_disponible"),
            Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
        ),
        total_validado=Coalesce(
            Sum("cantidad_validada"),
            Value(0, output_field=DecimalField(max_digits=20, decimal_places=6)),
        ),
    )

    context = {
        "sucursal": sucursal,
        "filas": filas,
        "total_disponible": agg["total_disponible"] or 0,
        "total_validado": agg["total_validado"] or 0,
        "q": q,
    }
    return render(request, "core/Movimientos/validar_stock_sucursal.html", context)


@login_required
def agregar_ubicacion_sucursal(request, sucursal_id):
    """
    Versión para SUCURSAL de 'agregar_ubicacion_bodega'.
    Crea una UbicacionSucursal asociada a la sucursal indicada.
    """
    sucursal = get_object_or_404(Sucursal, pk=sucursal_id)

    # instancia base con la sucursal ya seteada
    base_instance = UbicacionSucursal(sucursal=sucursal)

    if request.method == "POST":
        form = UbicacionSucursalForm(
            request.POST,
            instance=base_instance,              # <- importante
        )
        if form.is_valid():
            ubicacion = form.save(commit=False)

            # códigos que usas para armar el código completo
            area_cod    = form.cleaned_data.get("area_codigo")
            estante_cod = form.cleaned_data.get("estante_codigo")

            # si tu modelo UbicacionSucursal tiene este método (igual que bodega)
            if hasattr(ubicacion, "set_codigo"):
                ubicacion.set_codigo(area_cod, estante_cod)
            else:
                # fallback simple (ajusta al formato que uses)
                if area_cod and estante_cod:
                    # si quieres incluir el código de sucursal:
                    # ubicacion.codigo = f"{sucursal.codigo}-{area_cod}-{estante_cod}"
                    ubicacion.codigo = f"{area_cod}-{estante_cod}"

            ubicacion.save()
            return redirect("sucursal-list")   # o a donde quieras volver
    else:
        form = UbicacionSucursalForm(
            instance=base_instance,
            initial={
                # para que el hidden de la plantilla tenga valor
                "sucursal_codigo": getattr(sucursal, "codigo", ""),
            },
        )

    return render(
        request,
        "core/Movimientos/agregar_ubicacion_sucursal.html",
        {
            "form": form,
            "sucursal": sucursal,
        },
    )









import json
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.views.decorators.http import require_http_methods

from .models import Sucursal, Producto, Stock


@login_required
@require_http_methods(["GET", "POST"])
def escanear_qr_sucursal(request, sucursal_id):
    """
    Escanea QR (SKU). Si el producto existe en la sucursal -> redirige a validar-stock-sucursal.
    Si no existe -> responde con error para mostrar alerta.
    """
    sucursal = get_object_or_404(Sucursal, pk=sucursal_id)

    # ----- POST AJAX: validar SKU escaneado -----
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except json.JSONDecodeError:
            return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

        sku = (payload.get("sku") or "").strip()
        if not sku:
            return JsonResponse({"ok": False, "error": "SKU vacío."}, status=400)

        # 1) Verificar que el producto exista por SKU (ignorando mayúsculas/minúsculas)
        producto = Producto.objects.filter(sku__iexact=sku).first()
        if not producto:
            return JsonResponse(
                {"ok": False, "error": f"El producto con SKU {sku} no existe."},
                status=404,
            )

        # 2) Verificar que el producto tenga registro en la sucursal
        existe_en_sucursal = Stock.objects.filter(
            producto=producto,
            ubicacion_sucursal__sucursal=sucursal,
        ).exists()

        if not existe_en_sucursal:
            return JsonResponse(
                {
                    "ok": False,
                    "error": f"El producto {producto.sku} no existe en la sucursal {sucursal.nombre}.",
                },
                status=404,
            )

        # 3) URL correcta (OJO: nombre de la URL con GUIONES)
        redirect_url = (
            reverse("validar-stock-sucursal", args=[sucursal.id])
            + f"?producto={producto.id}"
        )

        return JsonResponse({"ok": True, "redirect_url": redirect_url})

    # ----- GET: mostrar visual de escaneo -----
    context = {
        "sucursal": sucursal,
    }
    return render(request, "core/Movimientos/escanear_qr_sucursal.html", context)

@login_required
@require_http_methods(["GET", "POST"])
def escanear_qr_bodega(request, bodega_id):
    """
    Escanea QR (SKU). Si el producto existe en la bodega -> redirige a validar-stock-bodega.
    Si no existe -> responde con error para mostrar alerta.
    """
    bodega = get_object_or_404(Bodega, pk=bodega_id)

    # ----- POST AJAX: validar SKU escaneado -----
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except json.JSONDecodeError:
            return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

        sku = (payload.get("sku") or "").strip()
        if not sku:
            return JsonResponse({"ok": False, "error": "SKU vacío."}, status=400)

        # 1) Verificar que el producto exista por SKU (ignorando mayúsculas/minúsculas)
        producto = Producto.objects.filter(sku__iexact=sku).first()
        if not producto:
            return JsonResponse(
                {"ok": False, "error": f"El producto con SKU {sku} no existe."},
                status=404,
            )

        # 2) Verificar que el producto tenga registro en la BODEGA
        existe_en_bodega = Stock.objects.filter(
            producto=producto,
            ubicacion_bodega__bodega=bodega,
        ).exists()

        if not existe_en_bodega:
            return JsonResponse(
                {
                    "ok": False,
                    "error": f"El producto {producto.sku} no existe en la bodega {bodega.nombre}.",
                },
                status=404,
            )

        # 3) URL correcta (OJO: nombre de la URL con GUIONES)
        redirect_url = (
            reverse("validar-stock-bodega", args=[bodega.id])
            + f"?producto={producto.id}"
        )

        return JsonResponse({"ok": True, "redirect_url": redirect_url})

    # ----- GET: mostrar visual de escaneo -----
    context = {
        "bodega": bodega,
    }
    return render(request, "core/Movimientos/escanear_qr_bodega.html", context)



def etiqueta_producto(request, pk): 
    p = get_object_or_404(Producto, pk=pk)

    # QR que contiene directamente el SKU del producto
    qr_sku = qr_url(p.sku, size="220x220")

    # Código de barras del SKU
    barcode = barcode_url(
        p.sku,
        bcid="code128",
        scale=4,
        height=14,
        includetext=True,
    )

    ctx = {
        "producto": p,
        "qr": qr_sku,          # QR del SKU
        "barcode": barcode,    # Código de barras del SKU
    }
    return render(request, "core/etiqueta_producto.html", ctx)

@login_required
@user_passes_test(_is_auditor)
def crear_orden_compra(request):
    """
    Crea una nueva Orden de Compra usando OrdenCompraForm.
    Si el formulario es válido hace form.save() y redirige.
    Si NO es válido, vuelve a mostrar el formulario con errores.
    """
    if request.method == "POST":
        form = OrdenCompraForm(request.POST)
        if form.is_valid():
            oc = form.save()
            messages.success(
                request,
                f"Orden de compra {oc.numero_orden} creada correctamente."
            )
            # Puedes cambiar 'finanzas_reporte' por donde quieras volver
            return redirect("finanzas_reporte")
        else:
            # Para depurar: puedes ver en consola qué falló
            print("❌ OrdenCompraForm errors:", form.errors)
    else:
        form = OrdenCompraForm()

    return render(request, "core/orden_compra_form.html", {"form": form})

@login_required
@user_passes_test(_is_auditor)
def crear_factura_proveedor(request):
    if request.method == "POST":
        form = FacturaProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('finanzas_reporte')  # Redirigir a la página de reporte
    else:
        form = FacturaProveedorForm()

    return render(request, "core/crear_factura_proveedor.html", {"form": form})

@login_required
@transaction.atomic
def crear_recepcion(request):
    if request.method == "POST":
        form = RecepcionMercaderiaForm(request.POST)
        if form.is_valid():
            recepcion = form.save(commit=False)
            # No establezcas manualmente 'recibido_en' si no es editable
            recepcion.save()
            messages.success(request, "Recepción creada correctamente.")
            return redirect('finanzas_reporte')  # O la URL que desees
    else:
        form = RecepcionMercaderiaForm()

    return render(request, "core/crear_recepcion.html", {"form": form})

@login_required
@user_passes_test(_is_auditor)
def crear_producto(request):
    if request.method == "POST":
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('finanzas_reporte')  # Redirigir a la página de reporte
    else:
        form = ProductoForm()

    return render(request, "core/crear_producto.html", {"form": form})


















































@login_required
def base_panel_control(request):
    # Obtener sucursales y bodegas
    sucursales = Sucursal.objects.filter(activo=True)
    bodegas = Bodega.objects.all()

    # Obtener productos con stock bajo
    stock_bajo = Stock.objects.filter(cantidad_disponible__lt=10)
    alertas_stock = stock_bajo.count()  # Número de productos con stock bajo

    # Sumar cantidades de productos por sucursal y bodega
    sucursales_activas = sucursales.count()
    bodegas_totales = bodegas.count()

    # Pasa estos datos al template
    context = {
        'sucursales': sucursales,
        'bodegas': bodegas,
        'alertas_stock': alertas_stock,
        'sucursales_activas': sucursales_activas,
        'bodegas_totales': bodegas_totales,
    }

    return render(request, 'core/base_panel_control.html', context)













# views.py
from django.shortcuts import render, redirect
from .forms import MarcaForm, UnidadMedidaForm, TasaImpuestoForm, CategoriaProductoForm

def centro_catalogo(request):

    marca_form = MarcaForm(prefix="marca")
    unidad_form = UnidadMedidaForm(prefix="unidad")
    tasa_form = TasaImpuestoForm(prefix="tasa")
    categoria_form = CategoriaProductoForm(prefix="categoria")

    if request.method == "POST":

        # Marca
        if "submit_marca" in request.POST:
            marca_form = MarcaForm(request.POST, prefix="marca")
            if marca_form.is_valid():
                marca_form.save()
                return redirect("centro-catalogo")

        # Unidad
        if "submit_unidad" in request.POST:
            unidad_form = UnidadMedidaForm(request.POST, prefix="unidad")
            if unidad_form.is_valid():
                unidad_form.save()
                return redirect("centro-catalogo")

        # Tasa
        if "submit_tasa" in request.POST:
            tasa_form = TasaImpuestoForm(request.POST, prefix="tasa")
            if tasa_form.is_valid():
                tasa_form.save()
                return redirect("centro-catalogo")

        # Categoría
        if "submit_categoria" in request.POST:
            categoria_form = CategoriaProductoForm(request.POST, prefix="categoria")
            if categoria_form.is_valid():
                categoria_form.save()
                return redirect("centro-catalogo")

    return render(request, "core/centro_catalogo.html", {
        "marca_form": marca_form,
        "unidad_form": unidad_form,
        "tasa_form": tasa_form,
        "categoria_form": categoria_form,
    })
